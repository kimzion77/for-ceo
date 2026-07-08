"""노무 가이드 DB API — 영세사업주 꿀팁.

자율점검 본질 (사업주 자가 점검) 기준 — `excluded_from_service=1` 항목은
모든 응답에서 자동 제외. audience 기본 = 'employer' (또는 'both').

설계 원칙
- 분쟁·신고·구제 류는 시드 자체에서 제외됐고, 추가 안전망으로 API 도 필터.
- 캐싱 — 가이드는 정적 데이터, ETag 또는 Cache-Control 길게.
- 룰엔진 위반 코드(V003 등) 와 cross-link — `wage_calc_formula` 활용.
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse, StreamingResponse

from cgr import db as _db
from cgr.api.auth import require_api_key


router = APIRouter(prefix="/guide", tags=["guide"])

# 양식 파일 저장소 — `cgr.db.get_db_path` 와 같은 ROOT 기준
_FORMS_DIR = Path(__file__).resolve().parents[3] / "data" / "forms"


def _query_all(sql: str, params: tuple = ()) -> list[dict[str, Any]]:
    with _db.connect() as conn:
        cur = conn.execute(sql, params)
        cols = [d[0] for d in cur.description] if cur.description else []
        return [dict(zip(cols, list(r))) for r in cur.fetchall()]


# ─────────────────────────────────────────────
# 1) GET /api/v1/guide/items — FAQ 항목
# ─────────────────────────────────────────────
@router.get(
    "/items",
    summary="가이드 FAQ 항목 (1·2·3 시트 통합)",
    description=(
        "audience: employer / worker / both. 기본 사업주 + 공통.\n"
        "category 로 필터 가능 (임금 검증 / 의무 경계 / 취업규칙 등)."
    ),
    dependencies=[Depends(require_api_key)],
)
def get_guide_items(
    audience: str | None = Query(
        default=None, description="필터 — employer / worker / both"
    ),
    category: str | None = Query(default=None, description="카테고리 부분 일치"),
):
    sql = (
        "SELECT code, audience, category, title, worker_reason, employer_reason, "
        "       key_points, related_laws, priority, applies_under_5, note "
        "FROM guide_item WHERE excluded_from_service = 0 "
    )
    params: list[Any] = []
    if audience:
        sql += "AND audience = ? "
        params.append(audience)
    else:
        # 기본 — 사업주 + 공통
        sql += "AND audience IN ('employer', 'both') "
    if category:
        sql += "AND category LIKE ? "
        params.append(f"%{category}%")
    sql += "ORDER BY priority, code"
    return {"items": _query_all(sql, tuple(params))}


@router.get(
    "/items/{code}",
    summary="가이드 항목 단건",
    dependencies=[Depends(require_api_key)],
)
def get_guide_item(code: str):
    rows = _query_all(
        "SELECT * FROM guide_item WHERE code = ? AND excluded_from_service = 0",
        (code,),
    )
    if not rows:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="not found")
    return rows[0]


# ─────────────────────────────────────────────
# 2) GET /api/v1/guide/glossary — 용어 사전
# ─────────────────────────────────────────────
@router.get(
    "/glossary",
    summary="용어 사전 (통상임금 vs 평균임금 등)",
    dependencies=[Depends(require_api_key)],
)
def get_glossary():
    return {
        "items": _query_all(
            "SELECT code, term, short_def, full_def, confusable_with, legal_basis "
            "FROM guide_glossary ORDER BY code"
        )
    }


# ─────────────────────────────────────────────
# 3) GET /api/v1/guide/by-size/{min_size} — 규모별 의무
# ─────────────────────────────────────────────
@router.get(
    "/by-size/{min_size}",
    summary="사업장 규모별 의무 — 홈 페이지에서 prefetch",
    description="min_size 예: '1인 이상', '5인 이상', '10인 이상', '30인 이상', '50인 이상'",
    dependencies=[Depends(require_api_key)],
)
def get_duties_by_size(min_size: str):
    # 정확 일치 + 누적 (예: 10인 이상은 1·5·10 모두 적용)
    SIZE_RANK = {
        "1인 이상": 1,
        "5인 이상": 5,
        "10인 이상": 10,
        "30인 이상": 30,
        "50인 이상": 50,
    }
    target_rank = SIZE_RANK.get(min_size, 0)
    if not target_rank:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"min_size must be one of {list(SIZE_RANK.keys())}",
        )
    rows = _query_all(
        "SELECT code, min_size, duty, description, related_docs, legal_basis "
        "FROM size_threshold_duty ORDER BY code"
    )
    # 본인 규모 이하 + 일치 모두 노출 (1인+ 의무는 5인+ 사업장에도 해당)
    applicable = [
        r for r in rows if SIZE_RANK.get(r["min_size"], 999) <= target_rank
    ]
    return {"size": min_size, "rank": target_rank, "duties": applicable}


# ─────────────────────────────────────────────
# 4) GET /api/v1/guide/by-stage/{stage} — 시기별 의무
# ─────────────────────────────────────────────
@router.get(
    "/by-stage/{stage}",
    summary="시기별 의무 (사업개시·채용·근로중·종료)",
    dependencies=[Depends(require_api_key)],
)
def get_duties_by_stage(stage: str):
    rows = _query_all(
        "SELECT code, stage, duty, description, deadline, legal_basis, priority, penalty "
        "FROM obligation_timeline "
        "WHERE excluded_from_service = 0 AND stage LIKE ? "
        "ORDER BY priority, code",
        (f"%{stage}%",),
    )
    return {"stage": stage, "duties": rows}


# ─────────────────────────────────────────────
# 4-b) GET /api/v1/guide/timeline — 전체 단계별 의무 (28개)
# ─────────────────────────────────────────────
@router.get(
    "/timeline",
    summary="전체 단계별 의무 (8단계 × 28항목)",
    description="사업 흐름(개시→채용→임금→…→종료) 순서대로 기한·우선순위·과태료 포함.",
    dependencies=[Depends(require_api_key)],
)
def get_timeline_all():
    rows = _query_all(
        "SELECT code, stage, duty, description, deadline, legal_basis, priority, penalty "
        "FROM obligation_timeline "
        "WHERE excluded_from_service = 0 "
        "ORDER BY stage, priority, code",
    )
    return {"items": rows}


# ─────────────────────────────────────────────
# 5) GET /api/v1/guide/forms — 신청 서식 (사업주용만)
# ─────────────────────────────────────────────
@router.get(
    "/forms",
    summary="신청 서식 카탈로그 (사업주 작성·제출용)",
    description="진정·구제 신청서 등은 시드에서 제외됨.",
    dependencies=[Depends(require_api_key)],
)
def get_forms(
    category: str | None = Query(default=None),
    audience: str | None = Query(default=None, description="명시 시 그 audience 만"),
):
    """양식 카탈로그 — 분쟁 양식은 시드에서 제외(EXCLUDED_FORM_CODES)되었고,
    남은 양식은 모두 정상 노무행정 양식이라 사업주가 모두 알아둬야 한다:

      - 근로자가 신청하는 출산·육아·산재 급여 → 사업주 확인서·협조 의무
      - 사업주가 발급하는 이직확인서·임금명세서
      - 사업주가 작성·신고하는 근로계약서·취업규칙

    그래서 default 는 audience 무관 — `excluded_from_service = 0` 만.
    """
    sql = (
        "SELECT code, category, form_name, purpose, submitter, submit_to, "
        "       submit_method, deadline, legal_basis, download_url, audience, "
        "       local_filename, local_mime, local_size, fetched_at "
        "FROM form_template WHERE excluded_from_service = 0 "
    )
    params: list[Any] = []
    if audience:
        sql += "AND audience = ? "
        params.append(audience)
    if category:
        sql += "AND category LIKE ? "
        params.append(f"%{category}%")
    sql += "ORDER BY code"
    items = _query_all(sql, tuple(params))
    # 클라이언트가 "직접 다운로드 가능" 한지 즉시 알 수 있도록 has_local 플래그 추가.
    # local_filename 이 있고 실제 파일도 있어야 true — 시드만 되고 파일 빠진 경우 false.
    for it in items:
        fn = it.get("local_filename")
        it["has_local"] = bool(fn and (_FORMS_DIR / fn).is_file())
    return {"items": items}


# ─────────────────────────────────────────────
# 5-b) GET /api/v1/guide/forms/{code}/download — 양식 파일 직접 다운로드
# ─────────────────────────────────────────────
@router.get(
    "/forms/{code}/download",
    summary="양식 파일 다운로드 — 로컬 파일이 있으면 스트림, 없으면 외부 URL 로 302",
    description=(
        "1) `local_filename` 컬럼이 채워져 있고 `backend/data/forms/<filename>` 파일이 "
        "실재하면 FileResponse 로 stream (정확한 MIME + Content-Disposition).\n"
        "2) 둘 다 없으면 `download_url` (고용노동부 자료실 등) 로 302 redirect — "
        "사용자가 외부 사이트에서라도 양식을 찾을 수 있게 폴백."
    ),
    dependencies=[Depends(require_api_key)],
)
def download_form(code: str):
    rows = _query_all(
        "SELECT code, form_name, local_filename, local_mime, download_url, "
        "excluded_from_service "
        "FROM form_template WHERE code = ?",
        (code,),
    )
    if not rows:
        raise HTTPException(status_code=404, detail=f"양식 코드 {code} 를 찾을 수 없습니다.")
    row = rows[0]
    if row["excluded_from_service"]:
        raise HTTPException(
            status_code=410,
            detail="이 양식은 자율점검 범위에서 제외됐습니다(분쟁·구제 신청서).",
        )

    local_filename = row["local_filename"]
    if local_filename:
        # 경로조작 방어(국정원 8대/KISA '경로 조작 및 자원 삽입') — DB 값이라도
        # ../ 등으로 _FORMS_DIR 밖을 가리키지 못하게 basename + 컨테인먼트 이중 확인.
        safe_name = Path(local_filename).name
        path = (_FORMS_DIR / safe_name).resolve()
        forms_root = _FORMS_DIR.resolve()
        if forms_root in path.parents and path.is_file():
            # Content-Disposition 헤더 — starlette/HTTP 표준이 latin-1 강제하므로
            # 한국어 파일명을 그대로 넣으면 UnicodeEncodeError. 해결:
            #   1) ASCII fallback (코드 기반) — 모든 브라우저가 인식
            #   2) RFC 5987 `filename*=UTF-8''<percent-encoded>` — 한국어 그대로 노출
            # 최신 브라우저는 filename* 우선 → 한국어 파일명으로 저장됨.
            # `quote` 가 ASCII 외 문자만 percent-encode 하므로 영문 파일명은 그대로.
            ext = Path(safe_name).suffix or ".bin"
            ascii_fallback = f"{row['code']}{ext}"
            cd = (
                f"attachment; filename=\"{ascii_fallback}\"; "
                f"filename*=UTF-8''{quote(safe_name)}"
            )
            return FileResponse(
                path,
                media_type=row["local_mime"] or "application/octet-stream",
                # FileResponse 의 filename= 인자도 latin-1 강제하므로 ASCII 만.
                # 진짜 파일명은 위 cd 헤더의 filename* 로 전달.
                filename=ascii_fallback,
                headers={"Content-Disposition": cd},
            )
        # local_filename 설정됐지만 실제 파일 없음 — 로깅용 경고는 운영자 몫
    # 폴백 — 외부 정부 사이트 URL
    if row["download_url"]:
        return RedirectResponse(url=row["download_url"], status_code=302)
    raise HTTPException(
        status_code=404,
        detail="이 양식의 다운로드 파일도, 외부 URL 도 등록돼 있지 않습니다.",
    )


# ─────────────────────────────────────────────
# 6) GET /api/v1/guide/wage-calc — 계산 공식 (V003~V006 연계)
# ─────────────────────────────────────────────
@router.get(
    "/wage-calc",
    summary="임금·수당 계산 공식 카탈로그",
    description=(
        "결과 페이지에서 V003 연장근로수당 finding 옆에 'CAL001 공식 보기' 같이 활용."
    ),
    dependencies=[Depends(require_api_key)],
)
def get_wage_calc(
    violation_code: str | None = Query(default=None),
):
    sql = (
        "SELECT code, category, calc_name, formula, conditions, limits, "
        "       legal_basis, note, related_violation_code "
        "FROM wage_calc_formula "
    )
    params: list[Any] = []
    if violation_code:
        sql += "WHERE related_violation_code = ? "
        params.append(violation_code)
    sql += "ORDER BY code"
    return {"items": _query_all(sql, tuple(params))}


# ─────────────────────────────────────────────
# 7) GET /api/v1/guide/orgs — 관할 기관 (분쟁 기관 제외)
# ─────────────────────────────────────────────
@router.get(
    "/orgs",
    summary="관할 기관 (사업주 이용 채널만)",
    dependencies=[Depends(require_api_key)],
)
def get_orgs():
    return {
        "items": _query_all(
            "SELECT code, org_class, org_name, duties, common_cases, phone, "
            "       online_channel, jurisdiction, note "
            "FROM gov_org WHERE excluded_from_service = 0 ORDER BY code"
        )
    }


# ─────────────────────────────────────────────
# 8) GET /api/v1/guide/audit — 근로감독 가이드
# ─────────────────────────────────────────────
@router.get(
    "/audit",
    summary="근로감독 종류·진행 절차",
    dependencies=[Depends(require_api_key)],
)
def get_audit_guide():
    return {
        "types": _query_all(
            "SELECT code, name, description, period_covered, legal_basis "
            "FROM audit_guide WHERE kind = 'type' ORDER BY code"
        ),
        "procedure": _query_all(
            "SELECT code, name, step_no, timing, description, legal_basis "
            "FROM audit_guide WHERE kind = 'procedure' "
            "ORDER BY step_no, code"
        ),
    }


# ─────────────────────────────────────────────
# 9) GET /api/v1/guide/required-docs — 비치 서류
# ─────────────────────────────────────────────
@router.get(
    "/required-docs",
    summary="법령상 의무 비치 서류",
    dependencies=[Depends(require_api_key)],
)
def get_required_docs():
    return {
        "items": _query_all(
            "SELECT code, classification, doc_name, description, "
            "       prep_time, retention_period, legal_basis, penalty "
            "FROM required_document ORDER BY classification, code"
        )
    }


# ─────────────────────────────────────────────
# 10) GET /api/v1/guide/lifecycle — 채용~종료 라이프사이클
# ─────────────────────────────────────────────
@router.get(
    "/lifecycle",
    summary="채용부터 종료까지 종합 가이드",
    dependencies=[Depends(require_api_key)],
)
def get_lifecycle():
    return {
        "items": _query_all(
            "SELECT code, phase, sub_topic, requirement, related_docs, "
            "       timing, legal_basis, note "
            "FROM employment_lifecycle ORDER BY code"
        )
    }


# ─────────────────────────────────────────────
# 11) GET /api/v1/guide/recruit — 채용 절차 준수사항
# ─────────────────────────────────────────────
@router.get(
    "/recruit",
    summary="채용 절차 준수사항 (채용공고~합격통지)",
    dependencies=[Depends(require_api_key)],
)
def get_recruit():
    return {
        "items": _query_all(
            "SELECT code, stage, duty, description, violation_examples, "
            "       penalty, applies_to, legal_basis, checkpoint "
            "FROM recruit_compliance ORDER BY code"
        )
    }


# ─────────────────────────────────────────────
# 12) GET /api/v1/guide/overview — 대시보드 한 화면 요약
# ─────────────────────────────────────────────
@router.get(
    "/overview",
    summary="가이드 전체 한 줄 요약 (사업주 대시보드용)",
    dependencies=[Depends(require_api_key)],
)
def get_overview():
    """프론트 /guide 페이지의 첫 진입 시 카운트만 빠르게."""
    return JSONResponse(
        {
            "guide_items": _query_all(
                "SELECT COUNT(*) AS n FROM guide_item "
                "WHERE excluded_from_service = 0 AND audience IN ('employer', 'both')"
            )[0]["n"],
            "obligations": _query_all(
                "SELECT COUNT(*) AS n FROM obligation_timeline WHERE excluded_from_service = 0"
            )[0]["n"],
            "wage_formulas": _query_all(
                "SELECT COUNT(*) AS n FROM wage_calc_formula"
            )[0]["n"],
            "glossary": _query_all("SELECT COUNT(*) AS n FROM guide_glossary")[0]["n"],
            "forms": _query_all(
                "SELECT COUNT(*) AS n FROM form_template "
                "WHERE excluded_from_service = 0 AND audience IN ('employer', 'both')"
            )[0]["n"],
            "orgs": _query_all(
                "SELECT COUNT(*) AS n FROM gov_org WHERE excluded_from_service = 0"
            )[0]["n"],
            "required_docs": _query_all(
                "SELECT COUNT(*) AS n FROM required_document"
            )[0]["n"],
            "lifecycle_steps": _query_all(
                "SELECT COUNT(*) AS n FROM employment_lifecycle"
            )[0]["n"],
        }
    )


# ═══════════════════════════════════════════════════════════════
# 13) POST /api/v1/guide/chat — 노무 가이드 챗봇
#
# 가이드 DB 정리된 자료(시기별 의무·규모별 의무·용어·기관·비치서류·라이프사이클)를
# 키워드 매칭으로 검색해 컨텍스트로 묶고, OpenAI LLM 으로 친근한 답변 생성.
#
# 결정성: temperature=0 + llm_cache 캐싱 (system+user 해시).
# ═══════════════════════════════════════════════════════════════

from pydantic import BaseModel, Field  # noqa: E402


class GuideChatTurn(BaseModel):
    role: str = Field(..., description="user 또는 assistant")
    content: str


class GuideChatIn(BaseModel):
    message: str = Field(..., min_length=1, max_length=2000)
    history: list[GuideChatTurn] | None = None


class RelatedFormHint(BaseModel):
    code: str
    form_name: str
    category: str
    audience: str
    has_local: bool
    purpose: str = ""


class GuideChatOut(BaseModel):
    answer: str
    matched_sources: list[str] = Field(
        default_factory=list,
        description="컨텍스트로 사용된 가이드 카테고리 (사용자 신뢰용)",
    )
    follow_ups: list[str] = Field(
        default_factory=list,
        description="이어서 물어볼만한 후속 질문 (사업주 자율점검 범위, 답변 컨텍스트 인지)",
    )
    related_forms: list[RelatedFormHint] = Field(
        default_factory=list,
        description="질문·답변에 등장한 주제와 관련된 사업주용 서식 (다운로드 chip 노출)",
    )
    clarify: str | None = Field(
        default=None,
        description="여러 변형 서식이 매칭된 경우 사용자에게 한 번 더 묻는 질문 (예: '어떤 유형의 근로계약서인가요?'). 없으면 null.",
    )


# ────────────────────────────────────────────────────────────
# 주제 → 서식 코드 매핑.
# 키는 정규식 (한국어 lowercase 비교). 값은 (form_codes, family_label, doc_topic)
# family_label 이 있으면 같은 family 안에서 여러 형이 매칭될 때 clarify 질문 생성.
# doc_topic=True 는 주제 자체가 '문서 이름'(근로계약서 등) — 질문에 등장하면
# 그 자체로 서식 수요로 본다. False 는 제도·급여 주제 — 질문에 서식·신청
# intent 가 함께 있을 때만 서식 chip 을 노출한다.
# ────────────────────────────────────────────────────────────
_TOPIC_TO_FORMS: list[tuple[str, list[str], str | None, bool]] = [
    # 근로계약서 family — 5종 변형 (정규/기간제/단시간/연소/건설일용) + 외국인
    (r"근로계약서|근로\s*계약", ["FRM001", "FRM002", "FRM003", "FRM004", "FRM005", "FRM030"], "근로계약서", True),
    # 취업규칙
    (r"취업규칙", ["FRM029", "FRM033"], None, True),
    # 임금명세서·임금대장
    (r"임금명세서|임금\s*명세|임금대장", ["FRM031", "FRM032"], None, True),
    # 4대보험
    (r"4\s*대\s*보험|사회보험|국민연금|건강보험|고용보험|산재보험", ["FRM006", "FRM007", "FRM008"], None, False),
    # 출산·육아·배우자 출산휴가
    (r"출산전후휴가|출산\s*휴가", ["FRM009"], None, False),
    (r"육아휴직|육아\s*휴직", ["FRM010", "FRM011"], None, False),
    (r"육아기\s*근로시간\s*단축", ["FRM012"], None, False),
    (r"배우자\s*출산", ["FRM013"], None, False),
    (r"고용안정장려금|출산육아기\s*고용안정", ["FRM014"], None, False),
    # 실업급여·이직
    (r"실업급여|수급자격", ["FRM015", "FRM017"], None, False),
    (r"이직확인서|이직\s*확인", ["FRM016"], None, True),
    # 퇴직금·퇴직연금
    (r"퇴직금|퇴직\s*급여", ["FRM018"], None, False),
    (r"퇴직연금|db\s*형|dc\s*형|확정급여형|확정기여형", ["FRM019", "FRM034"], "퇴직연금", False),
    # 산재
    (r"산재|업무상\s*재해|요양급여|휴업급여|장해급여", ["FRM025", "FRM026", "FRM027", "FRM028"], "산재", False),
    # 외국인
    (r"외국인", ["FRM030"], None, False),
]

# 질문에서 '서식이 필요하다'는 의도를 나타내는 표현 — 제도 주제(doc_topic=False)는
# 이 intent 가 질문에 함께 있을 때만 서식 chip 노출.
_FORM_INTENT_RE = (
    r"서식|양식|신청|신고|확인서|증명서|규약|서류|"
    r"작성|제출|다운로드|받고\s*싶|받을\s*수|받아\s*보|어디서\s*받"
)


def _detect_related_forms(question: str, answer: str) -> tuple[list[RelatedFormHint], str | None]:
    """질문 텍스트를 스캔해 관련 서식 코드를 수집.

    답변 텍스트는 보지 않는다 — 노무 답변에는 '근로계약'·'고용보험' 같은 주제어가
    거의 항상 등장해, 답변까지 스캔하면 사실상 모든 질문에 서식이 떴다.
    문서 이름 주제(doc_topic)는 질문 등장만으로, 제도 주제는 질문에 서식·신청
    intent 가 함께 있을 때만 노출한다.

    같은 family 안에서 2개 이상 매칭되면 clarify 질문 생성.
    """
    import re as _re

    del answer  # 의도적으로 미사용 — docstring 참조
    text = question.lower().replace(" ", "")
    has_intent = bool(_re.search(_FORM_INTENT_RE.replace(r"\s*", ""), text))
    matched_codes: list[str] = []
    matched_families: set[str] = set()
    for pat, codes, family, doc_topic in _TOPIC_TO_FORMS:
        # 패턴은 공백 제거된 텍스트에 매칭하기 위해 \s* 제거
        if not _re.search(pat.replace(r"\s*", ""), text):
            continue
        if not (doc_topic or has_intent):
            continue
        for c in codes:
            if c not in matched_codes:
                matched_codes.append(c)
        if family:
            matched_families.add(family)

    if not matched_codes:
        return [], None

    # 매칭된 코드로 form_template 조회 — 사업주(employer) 또는 both 만
    placeholders = ",".join(["?"] * len(matched_codes))
    rows = _query_all(
        f"SELECT code, form_name, category, audience, purpose, local_filename, download_url "
        f"FROM form_template "
        f"WHERE code IN ({placeholders}) AND audience IN ('employer', 'both') "
        f"ORDER BY code",
        tuple(matched_codes),
    )
    hints: list[RelatedFormHint] = []
    for r in rows:
        hints.append(RelatedFormHint(
            code=r["code"],
            form_name=r["form_name"],
            category=r["category"],
            audience=r["audience"],
            has_local=bool(r.get("local_filename")),
            purpose=(r.get("purpose") or "")[:100],
        ))

    # clarify: 같은 family 에서 2개 이상 매칭된 경우 사용자에게 한 번 더 묻기
    clarify: str | None = None
    family_counts: dict[str, int] = {}
    for h in hints:
        for pat, codes, family, _doc_topic in _TOPIC_TO_FORMS:
            if family and h.code in codes:
                family_counts[family] = family_counts.get(family, 0) + 1
    for family, cnt in family_counts.items():
        if cnt >= 2:
            if family == "근로계약서":
                clarify = "근로계약서는 근로자 유형별로 양식이 달라요. 어떤 유형인가요? 아래 버튼에서 골라 받으세요."
            elif family == "퇴직연금":
                clarify = "퇴직연금은 제도 유형(DB형/DC형)에 따라 표준규약이 달라요. 운영하시는 제도 기준으로 골라 받으세요."
            elif family == "산재":
                clarify = "산재 관련 서식은 신청 단계별로 양식이 달라요. 필요한 단계의 서식을 골라 받으세요."
            break

    return hints, clarify


def _search_guide_context(query: str, *, per_table: int = 4) -> tuple[str, list[str]]:
    """사용자 질문 키워드로 가이드 DB 검색 → LLM 컨텍스트 텍스트 + 사용된 카테고리.

    각 테이블에서 LIKE 검색으로 상위 N건만 추려 토큰 절약.
    """
    if not query:
        return "", []
    # 한국어 키워드 추출 — 단순 split + 2자 이상
    tokens = [t.strip() for t in query.replace('?', ' ').split() if len(t.strip()) >= 2]
    # 조사 변형 보정 — "근로감독시"·"감독시" 처럼 끝에 조사가 붙으면 LIKE 매칭이
    # 안 돼 검색을 놓친다. 끝 1글자가 흔한 조사면 떼어낸 형태도 검색어에 추가.
    _PARTICLES = set("은는이가을를에의도만과와로서시께란")
    _extra: list[str] = []
    for t in tokens:
        if len(t) >= 3 and t[-1] in _PARTICLES:
            _extra.append(t[:-1])
        if len(t) >= 4 and t[-2:] in ("에서", "으로", "에게", "이나", "라고", "부터", "까지"):
            _extra.append(t[:-2])
    tokens = list(dict.fromkeys([*tokens, *_extra]))  # 순서 유지 dedupe
    if not tokens:
        return "", []
    sources: list[str] = []
    blocks: list[str] = []

    def _like_search(sql_select: str, text_cols: list[str], label: str, max_n: int = per_table):
        """주어진 SELECT 에 LIKE 조건 추가해 매칭 행 반환."""
        if not text_cols:
            return []
        # 모든 토큰을 OR 로 — 어느 컬럼이든 어느 토큰이든 매칭되면 hit
        like_clauses = []
        params: list[Any] = []
        for tk in tokens:
            for col in text_cols:
                like_clauses.append(f"{col} LIKE ?")
                params.append(f"%{tk}%")
        where = " OR ".join(like_clauses)
        # sql_select 에 이미 WHERE 가 있으면 AND 로 이어붙임 — 'WHERE … WHERE …'
        # SQL 오류로 해당 테이블 검색이 조용히 빈 결과가 되던 버그 수정.
        joiner = "AND" if " where " in sql_select.lower() else "WHERE"
        full_sql = f"{sql_select} {joiner} ({where}) LIMIT {max_n}"
        try:
            return _query_all(full_sql, tuple(params))
        except Exception:
            return []

    # 1) 시기별 의무
    rows = _like_search(
        "SELECT stage, duty, description, deadline, legal_basis "
        "FROM obligation_timeline WHERE excluded_from_service = 0 ",
        ["stage", "duty", "description", "legal_basis"],
        "obligation",
    )
    if rows:
        sources.append("시기별 의무")
        lines = ["[시기별 의무]"]
        for r in rows:
            lines.append(
                f"- [{r['stage']}] {r['duty']}: {r.get('description','')} "
                f"(근거 {r.get('legal_basis','')}, 기한 {r.get('deadline','')})"
            )
        blocks.append("\n".join(lines))

    # 2) 사업장 규모별 의무
    rows = _like_search(
        "SELECT min_size, duty, description, legal_basis "
        "FROM size_threshold_duty ",
        ["min_size", "duty", "description", "legal_basis"],
        "size",
    )
    if rows:
        sources.append("규모별 의무")
        lines = ["[사업장 규모별 의무]"]
        for r in rows:
            lines.append(
                f"- [{r['min_size']}] {r['duty']}: {r.get('description','')} "
                f"(근거 {r.get('legal_basis','')})"
            )
        blocks.append("\n".join(lines))

    # 3) 용어 사전
    rows = _like_search(
        "SELECT term, short_def, full_def, confusable_with, legal_basis "
        "FROM guide_glossary ",
        ["term", "short_def", "full_def", "confusable_with"],
        "glossary",
    )
    if rows:
        sources.append("용어 사전")
        lines = ["[용어 사전]"]
        for r in rows:
            extra = f" 혼동: {r['confusable_with']}" if r.get("confusable_with") else ""
            lines.append(
                f"- {r['term']}: {r.get('short_def','')}. {r.get('full_def','')[:200]}{extra}"
            )
        blocks.append("\n".join(lines))

    # 4) 기관
    rows = _like_search(
        "SELECT org_name, org_class, duties, common_cases, phone, online_channel "
        "FROM gov_org WHERE excluded_from_service = 0 ",
        ["org_name", "duties", "common_cases", "org_class"],
        "org",
    )
    if rows:
        sources.append("정부 기관")
        lines = ["[정부 기관·온라인 채널]"]
        for r in rows:
            ch = r.get("online_channel") or r.get("phone") or ""
            lines.append(f"- {r['org_name']} ({r['org_class']}): {r.get('duties','')} · {ch}")
        blocks.append("\n".join(lines))

    # 5) 비치 서류
    rows = _like_search(
        "SELECT doc_name, classification, description, retention_period, legal_basis "
        "FROM required_document ",
        ["doc_name", "description", "classification", "legal_basis"],
        "doc",
    )
    if rows:
        sources.append("비치 서류")
        lines = ["[비치·보존 서류]"]
        for r in rows:
            lines.append(
                f"- {r['doc_name']} ({r['classification']}): {r.get('description','')} "
                f"보존 {r.get('retention_period','')} (근거 {r.get('legal_basis','')})"
            )
        blocks.append("\n".join(lines))

    # 6) 라이프사이클
    rows = _like_search(
        "SELECT phase, sub_topic, requirement, related_docs, legal_basis "
        "FROM employment_lifecycle ",
        ["phase", "sub_topic", "requirement", "legal_basis"],
        "lifecycle",
    )
    if rows:
        sources.append("고용 생애주기")
        lines = ["[고용 생애주기]"]
        for r in rows:
            lines.append(
                f"- [{r['phase']} / {r['sub_topic']}] {r['requirement']} "
                f"(서류 {r.get('related_docs','')}, 근거 {r.get('legal_basis','')})"
            )
        blocks.append("\n".join(lines))

    # 7) 채용 컴플라이언스
    rows = _like_search(
        "SELECT stage, duty, description, violation_examples, penalty, legal_basis "
        "FROM recruit_compliance ",
        ["stage", "duty", "description", "violation_examples"],
        "recruit",
    )
    if rows:
        sources.append("채용 컴플라이언스")
        lines = ["[채용 단계 준수사항]"]
        for r in rows:
            lines.append(
                f"- [{r['stage']}] {r['duty']}: {r.get('description','')} "
                f"위반사례 {r.get('violation_examples','')} (벌칙 {r.get('penalty','')})"
            )
        blocks.append("\n".join(lines))

    # 8) 가이드 FAQ (guide_item) — 노무제공자 공통 표준계약서 등 주제별 안내
    rows = _like_search(
        "SELECT category, title, key_points, related_laws, note "
        "FROM guide_item WHERE excluded_from_service = 0 ",
        ["category", "title", "key_points", "note"],
        "guide_item",
    )
    if rows:
        sources.append("가이드 FAQ")
        lines = ["[가이드 FAQ]"]
        for r in rows:
            lines.append(
                f"- [{r['category']}] {r['title']}: {r.get('key_points','')[:400]} "
                f"(근거 {r.get('related_laws','')})"
            )
        blocks.append("\n".join(lines))

    return "\n\n".join(blocks), sources


_GUIDE_CHAT_SYSTEM = (
    "너는 영세사업주를 위한 노동법 가이드 챗봇이다. 친근하고 명확한 톤으로 답하되,\n"
    "반드시 아래 [가이드 DB 컨텍스트] 의 정리된 자료를 1차 근거로 인용한다.\n\n"
    "원칙:\n"
    "1) 사업주가 알아야 할 의무·서식·절차·기관 정보 중심으로 답변. 분쟁·진정·구제는 안내 X.\n"
    "2) 답변 끝에 '관련 법령:' 한 줄로 근거 법령을 콤마로 묶어 표시.\n"
    "3) **공인노무사 상담 안내는 [정말 필요한 경우에만] 추가** — 매 답변마다 절대 붙이지 마라.\n"
    "   추가 조건 (이 중 하나라도 해당될 때만):\n"
    "   (a) 사실관계가 복잡하거나 판례가 갈리는 회색지대 (예: 정기상여금 통상임금성, 포괄임금제\n"
    "       유효성, 근로자성 판단, 부당해고 사유 정당성)\n"
    "   (b) 사업장 개별 사정에 따라 결론이 크게 달라지는 사안 (예: 단축근로 적용 범위, 취업규칙\n"
    "       불이익 변경 동의 요건)\n"
    "   (c) 가이드 DB 컨텍스트에 없는 영역이거나, LLM 일반지식만으로 답변한 경우\n"
    "   기본 의무·서식·신고 절차·기간·법령 인용 같은 명확한 사실 안내에는 노무사 권장 문구를\n"
    "   붙이지 마라. (관할 고용센터는 노무 상담 기관이 아님 — '고용센터 상담' 안내는 절대 금지.\n"
    "   단, 지원금·급여 신청 절차 안내는 고용센터 가능)\n"
    "4) 통상임금 판단은 2024.12.19 대법원 전원합의체 판결(2020다247190) 반영 — 고정성 요건 폐기,\n"
    "   소정근로 대가성·정기성·일률성 3요소만으로 판단.\n"
    "5) 답변은 2~5문장으로 간결. 필요하면 번호 목록 사용.\n"
    "6) 분쟁성 질문(진정·신고 등)이 들어오면 '본 서비스는 사업주 자율점검용입니다. 분쟁은\n"
    "   관할 지방고용노동청을 통해 진행해 주세요' 로 안내.\n"
    "7) **범위 밖 질문 거절** — 노동법·노무·사업장 운영(임금·근로시간·휴가·해고·보험·취업규칙·\n"
    "   근로계약서·임금명세서·노무제공자 계약·산재·출산육아·퇴직 등)과 무관한 질문은 답변하지\n"
    "   말고 다음 문구로 종결:\n"
    "     '죄송하지만 사업주 노무 관리 범위 밖 질문이라 답변드리기 어려워요. 노동법·근로조건·\n"
    "      보험·서식·계산 등 다른 노무 관련 질문이 있으시면 도와드릴게요.'\n"
    "   (예: '맛집 추천', '오늘 날씨', '주식 사는 법', '코딩 도와줘' 등 → 거절 문구만)\n"
    "   범위 밖이면 '관련 법령:' 줄과 [추천질문] 섹션 모두 출력 금지.\n"
    "8) **제재·처벌 표현 규칙(엄수)** — '전과', '벌금=형사처벌(전과)' 같은 표현은 절대 쓰지 마라.\n"
    "   위반 시 불이익은 반드시 '과태료·벌금 등 행정·형사 제재로 이어질 수 있습니다' 처럼\n"
    "   중립적으로 안내하고, 'A=B(전과)' 식 등식·낙인적 표현은 쓰지 않는다.\n\n"
    "[후속 추천 질문 — 반드시 출력]\n"
    "답변 본문 + '관련 법령' 표시 다음에 빈 줄 한 칸 후 정확히 다음 형식으로 추가:\n"
    "  [추천질문]\n"
    "  - <질문1>\n"
    "  - <질문2>\n"
    "  - <질문3>\n\n"
    "추천질문 작성 규칙:\n"
    "- 사용자가 방금 던진 질문과 다른 각도로, 답변 내용에서 자연스럽게 이어지는 후속 질문 3개.\n"
    "- 사업주 자율점검 범위(의무·서식·절차·기관·생애주기·채용·임금계산) 안에서만.\n"
    "- 짧고 구체적으로 (한 줄 ≤ 25자 권장).\n"
    "- 분쟁·진정·소송·노동위원회 관련은 추천하지 말 것.\n"
    "- 같은 주제 다른 측면(예: 의무 → 위반 시 제재, 신고 절차, 관련 서식, 상한·예외)으로 분산.\n"
    "예시:\n"
    "  [추천질문]\n"
    "  - 5인 미만은 어떻게 달라요?\n"
    "  - 위반 시 사업주 과태료는?\n"
    "  - 관련 표준 서식 어디서 받아요?"
)


_FOLLOWUP_BLOCK_RE = __import__('re').compile(
    r"\[추천질문\]\s*\n((?:\s*[-·●]\s*[^\n]+\n?)+)",
    flags=__import__('re').MULTILINE,
)


def _extract_followups(answer: str) -> tuple[str, list[str]]:
    """LLM 응답에서 [추천질문] 블록 분리 — 본문은 그 라인 제거 후 반환.

    포맷:
        ...본문...
        관련 법령: ...

        [추천질문]
        - 질문1
        - 질문2
        - 질문3
    """
    m = _FOLLOWUP_BLOCK_RE.search(answer)
    if not m:
        return answer.strip(), []
    block = m.group(1)
    lines = [
        ln.strip().lstrip("-·● ").strip()
        for ln in block.split("\n")
        if ln.strip()
    ]
    # 빈 항목·너무 짧은 항목 제거 + 최대 3개로 제한
    items = [ln for ln in lines if len(ln) >= 4][:3]
    # 본문에서 블록 제거
    body = answer[: m.start()].rstrip()
    return body, items


def _sanitize_sanctions(text: str) -> str:
    """제재·처벌 표현 결정적 정제.

    사용자 요구: '전과', '벌금=형사처벌(전과)' 같은 낙인적·등식형 표현 금지.
    위반 불이익은 '과태료·벌금 등 행정·형사 제재' 로 중립 안내.
    LLM 준수에 의존하지 않도록 (a) 가이드 DB 컨텍스트, (b) 최종 답변 본문 양쪽에
    적용한다. (같은 입력 → 같은 출력: 결정성 유지)
    """
    if not text:
        return text
    import re

    s = text
    for a, b in (
        ("벌금=형사처벌(전과)", "과태료·벌금 등 행정·형사 제재"),
        ("벌금 = 형사처벌(전과)", "과태료·벌금 등 행정·형사 제재"),
        ("형사처벌(전과)", "형사 제재"),
        ("벌금=형사처벌", "벌금 등 형사 제재"),
        ("벌금 = 형사처벌", "벌금 등 형사 제재"),
    ):
        s = s.replace(a, b)
    # 괄호 주석 '(전과...)' 제거
    s = re.sub(r"\s*\(\s*전과[^)]*\)", "", s)
    # 잔여 '전과(+조사)' 언급 제거 (낙인 표현 차단)
    s = re.sub(r"전과(가|는|을|를|로|기록)?", "", s)
    # 치환 흔적 정리 — 이중 공백 / 구두점 앞 공백
    s = re.sub(r"[ \t]{2,}", " ", s)
    s = re.sub(r"\s+([,.)])", r"\1", s)
    return s.strip()


@router.post(
    "/chat",
    response_model=GuideChatOut,
    summary="노무 가이드 챗봇 — 가이드 DB 컨텍스트 + LLM",
    dependencies=[Depends(require_api_key)],
)
def post_guide_chat(body: GuideChatIn, request: Request) -> GuideChatOut:
    msg = body.message.strip()
    if not msg:
        raise HTTPException(status_code=422, detail="질문이 비어있어요.")

    # 가이드 DB 검색 → LLM 컨텍스트 (제재 표현 정제 후 주입 — LLM이 '전과' 등을 echo 못 하게)
    ctx_text, sources = _search_guide_context(msg)
    ctx_text = _sanitize_sanctions(ctx_text)

    # 이전 대화 (최근 6턴)
    hist_lines: list[str] = []
    if body.history:
        for h in body.history[-6:]:
            role = h.role if h.role in ("user", "assistant") else "user"
            hist_lines.append(f"- {role}: {h.content[:600]}")

    user_prompt_parts: list[str] = []
    if ctx_text:
        user_prompt_parts.append("[가이드 DB 컨텍스트 — 1차 근거로 사용]\n" + ctx_text)
    else:
        user_prompt_parts.append(
            "[가이드 DB 컨텍스트] (사용자 질문과 직접 매칭되는 자료가 없음 — 일반 노동법 상식으로 답하되 마지막에 '관할 고용센터 상담 권장' 안내)"
        )
    if hist_lines:
        user_prompt_parts.append("[이전 대화]\n" + "\n".join(hist_lines))
    user_prompt_parts.append(f"[사용자 질문] {msg}")
    user_prompt = "\n\n".join(user_prompt_parts)

    # LLM 호출 — cgr.ec.services.chat 와 동일 패턴 (캐시 + 재시도)
    import time

    from openai import APIConnectionError, APITimeoutError, OpenAI, RateLimitError

    from cgr import llm_cache
    from cgr.config import get_api_key, get_llm_model

    model_name = get_llm_model()
    from cgr import prompt_store
    from cgr.upload_tracker import anon_visitor
    from cgr.store import analytics as _an

    # 관리자 override 가 있으면 그 프롬프트, 없으면 코드 기본값 (즉시 적용)
    system_prompt = prompt_store.get_or_default("guide_chat", _GUIDE_CHAT_SYSTEM)
    _visitor = anon_visitor(request)
    cache_key = llm_cache.make_key(
        system=system_prompt,
        user=user_prompt,
        schema={"kind": "guide_chat"},
        model=model_name,
    )
    cached = llm_cache.get(cache_key)
    if cached and isinstance(cached.get("text"), str):
        body, fups = _extract_followups(cached["text"])
        body = _sanitize_sanctions(body)
        _an.log_interaction(kind="챗봇", model=model_name, input_text=msg, output_text=body, visitor=_visitor)
        rel_forms, clarify = _detect_related_forms(msg, body)
        return GuideChatOut(
            answer=body,
            matched_sources=sources,
            follow_ups=fups,
            related_forms=rel_forms,
            clarify=clarify,
        )

    client = OpenAI(api_key=get_api_key(), timeout=60.0)
    last_err: Exception | None = None
    backoff = (2.0, 5.0, 10.0)
    for attempt in range(3):
        try:
            resp = client.chat.completions.create(
                model=model_name,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt},
                ],
                temperature=0,
                top_p=1,
            )
            text = (resp.choices[0].message.content or "").strip()
            if not text:
                raise RuntimeError("chat 응답이 비어 있습니다.")
            llm_cache.put(cache_key, {"text": text})
            body, fups = _extract_followups(text)
            body = _sanitize_sanctions(body)
            _an.log_interaction(kind="챗봇", model=model_name, input_text=msg, output_text=body, visitor=_visitor)
            rel_forms, clarify = _detect_related_forms(msg, body)
            return GuideChatOut(
                answer=body,
                matched_sources=sources,
                follow_ups=fups,
                related_forms=rel_forms,
                clarify=clarify,
            )
        except (APITimeoutError, APIConnectionError, RateLimitError) as e:
            last_err = e
            if attempt < 2:
                time.sleep(backoff[attempt])
                continue
            raise HTTPException(status_code=502, detail=f"LLM 호출 실패: {e}")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"chat 처리 실패: {e}")
    raise HTTPException(status_code=500, detail=f"chat 실패: {last_err}")


# ═══════════════════════════════════════════════════════════════
# 14) GET /api/v1/guide/export.xlsx — 가이드 데이터 통합 Excel 다운로드
#
# 6개 시트: 시기별 의무 / 규모별 의무 / 용어 사전 / 정부 기관 / 비치 서류 /
#           고용 생애주기 / 채용 컴플라이언스
# 사장님이 오프라인에서도 참고할 수 있도록 정리된 자료를 한 파일로.
# ═══════════════════════════════════════════════════════════════
@router.get(
    "/export.xlsx",
    summary="가이드 데이터 통합 Excel 다운로드",
    dependencies=[Depends(require_api_key)],
)
def export_guide_xlsx():
    """openpyxl 로 다중 시트 xlsx 생성 후 스트림."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=503, detail="openpyxl 미설치 — pip install openpyxl")

    wb = Workbook()
    # 헤더 스타일 — 옅은 브랜드 배경 + 굵은 글씨
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0B3D91")
    header_align = Alignment(vertical="center", horizontal="center")

    def _add_sheet(title: str, headers: list[str], rows: list[dict], col_widths: list[int]):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        for r in rows:
            ws.append([r.get(k, "") or "" for k in r.keys()])
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w
        ws.row_dimensions[1].height = 28
        # 본문 줄 wrap
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 기본 첫 시트 제거 후 순서대로 추가
    wb.remove(wb.active)

    # 1) 시기별 의무
    _add_sheet(
        "시기별 의무",
        ["코드", "단계", "의무", "설명", "기한", "근거 법령", "벌칙", "우선순위"],
        [
            {
                "code": r["code"],
                "stage": r["stage"],
                "duty": r["duty"],
                "description": r["description"],
                "deadline": r["deadline"],
                "legal_basis": r["legal_basis"],
                "penalty": r["penalty"],
                "priority": r["priority"],
            }
            for r in _query_all(
                "SELECT code, stage, duty, description, deadline, legal_basis, penalty, priority "
                "FROM obligation_timeline WHERE excluded_from_service = 0 "
                "ORDER BY stage, code"
            )
        ],
        [10, 12, 28, 50, 18, 28, 32, 10],
    )

    # 2) 규모별 의무
    _add_sheet(
        "규모별 의무",
        ["코드", "최소 규모", "의무", "설명", "관련 서류", "근거 법령"],
        [
            {
                "code": r["code"],
                "min_size": r["min_size"],
                "duty": r["duty"],
                "description": r["description"],
                "related_docs": r["related_docs"],
                "legal_basis": r["legal_basis"],
            }
            for r in _query_all(
                "SELECT code, min_size, duty, description, related_docs, legal_basis "
                "FROM size_threshold_duty ORDER BY "
                "CASE min_size WHEN '1인 이상' THEN 1 WHEN '5인 이상' THEN 5 "
                "  WHEN '10인 이상' THEN 10 WHEN '30인 이상' THEN 30 "
                "  WHEN '50인 이상' THEN 50 ELSE 999 END, code"
            )
        ],
        [10, 12, 30, 50, 25, 28],
    )

    # 3) 용어 사전
    _add_sheet(
        "용어 사전",
        ["코드", "용어", "짧은 정의", "상세 정의", "헷갈리는 용어", "근거"],
        [
            {
                "code": r["code"],
                "term": r["term"],
                "short_def": r["short_def"],
                "full_def": r["full_def"],
                "confusable_with": r["confusable_with"],
                "legal_basis": r["legal_basis"],
            }
            for r in _query_all(
                "SELECT code, term, short_def, full_def, confusable_with, legal_basis "
                "FROM guide_glossary ORDER BY code"
            )
        ],
        [10, 22, 50, 70, 30, 22],
    )

    # 4) 정부 기관
    _add_sheet(
        "정부 기관",
        ["코드", "기관 분류", "기관명", "담당 업무", "흔한 활용 사례", "전화", "온라인 채널", "관할"],
        [
            {
                "code": r["code"],
                "org_class": r["org_class"],
                "org_name": r["org_name"],
                "duties": r["duties"],
                "common_cases": r["common_cases"],
                "phone": r["phone"],
                "online_channel": r["online_channel"],
                "jurisdiction": r["jurisdiction"],
            }
            for r in _query_all(
                "SELECT code, org_class, org_name, duties, common_cases, phone, "
                "       online_channel, jurisdiction "
                "FROM gov_org WHERE excluded_from_service = 0 ORDER BY org_class, code"
            )
        ],
        [10, 18, 26, 40, 40, 14, 36, 22],
    )

    # 5) 비치 서류
    _add_sheet(
        "비치 서류",
        ["코드", "분류", "서류명", "설명", "작성 시기", "보존 기간", "근거 법령", "벌칙"],
        [
            {
                "code": r["code"],
                "classification": r["classification"],
                "doc_name": r["doc_name"],
                "description": r["description"],
                "prep_time": r["prep_time"],
                "retention_period": r["retention_period"],
                "legal_basis": r["legal_basis"],
                "penalty": r["penalty"],
            }
            for r in _query_all(
                "SELECT code, classification, doc_name, description, prep_time, "
                "       retention_period, legal_basis, penalty "
                "FROM required_document ORDER BY classification, code"
            )
        ],
        [10, 14, 30, 50, 18, 18, 28, 22],
    )

    # 6) 고용 생애주기
    _add_sheet(
        "고용 생애주기",
        ["코드", "단계", "세부 주제", "요건", "관련 서류", "시기", "근거"],
        [
            {
                "code": r["code"],
                "phase": r["phase"],
                "sub_topic": r["sub_topic"],
                "requirement": r["requirement"],
                "related_docs": r["related_docs"],
                "timing": r["timing"],
                "legal_basis": r["legal_basis"],
            }
            for r in _query_all(
                "SELECT code, phase, sub_topic, requirement, related_docs, timing, legal_basis "
                "FROM employment_lifecycle ORDER BY phase, code"
            )
        ],
        [10, 14, 22, 50, 22, 16, 24],
    )

    # 7) 채용 컴플라이언스
    _add_sheet(
        "채용 컴플라이언스",
        ["코드", "단계", "의무", "설명", "위반 사례", "벌칙", "적용 대상", "근거", "점검 포인트"],
        [
            {
                "code": r["code"],
                "stage": r["stage"],
                "duty": r["duty"],
                "description": r["description"],
                "violation_examples": r["violation_examples"],
                "penalty": r["penalty"],
                "applies_to": r["applies_to"],
                "legal_basis": r["legal_basis"],
                "checkpoint": r["checkpoint"],
            }
            for r in _query_all(
                "SELECT code, stage, duty, description, violation_examples, penalty, "
                "       applies_to, legal_basis, checkpoint "
                "FROM recruit_compliance ORDER BY stage, code"
            )
        ],
        [10, 14, 26, 45, 36, 22, 18, 22, 30],
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = "영세사업장_노무_가이드.xlsx"
    ascii_fallback = "labor-guide.xlsx"
    cd = (
        f"attachment; filename=\"{ascii_fallback}\"; "
        f"filename*=UTF-8''{quote(fname)}"
    )
    return StreamingResponse(
        buf,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": cd},
    )
