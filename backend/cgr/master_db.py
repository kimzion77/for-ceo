"""취업규칙 마스터 DB (xlsx) 로더 — 단일 진실원(SSoT).

- 경로: 2026판 우선 → 2025판 fallback
- 14개 컬럼: A 번호 / B 조제목 / C 필수선택 / D 본문 / E 착안사항 / F 참고
              / G 관련법령 / H 연관주제 / I 벌칙 / J 개정슬롯 / K 갱신 / L 구법
              / M 관련규정(빈출) / N 빈출지적
- 사용:
    db = get_master_db()
    db.title(33) -> "연차유급휴가"
    db.article(33) -> {title, scope, body, note, law, topic, penalty, ...}
"""
from __future__ import annotations

import os
from functools import lru_cache
from pathlib import Path
from typing import Any

# 컬럼 인덱스 (1-based, openpyxl 기준)
COLS = {
    "no": 1,         "title": 2,      "scope": 3,      "body": 4,
    "guide": 5,      "note": 6,       "law": 7,        "topic": 8,
    "penalty": 9,    "amend_slot": 10, "amend_new": 11, "amend_old": 12,
    "freq_clause": 13, "freq_issue": 14,
}

# 레포 루트 (backend/cgr/master_db.py → parents[2] = repo root).
# 배포 환경(Render/Linux)에서는 절대 Windows 경로가 없으므로 이 상대 경로가 1순위.
_REPO_ROOT = Path(__file__).resolve().parents[2]
_BACKEND_DATA = Path(__file__).resolve().parents[1] / "data"

# 후보 경로 (우선순위 순) — 상대경로(배포 호환) → Windows 로컬
DEFAULT_PATHS = [
    # ── 레포 상대 경로 (Render/Linux/모든 환경 공통) ──
    str(_REPO_ROOT / "취업규칙 마스터 db (2026).xlsx"),
    str(_REPO_ROOT / "취업규칙 마스터 db.xlsx"),
    str(_BACKEND_DATA / "취업규칙 마스터 db (2026).xlsx"),
    str(_BACKEND_DATA / "취업규칙 마스터 db.xlsx"),
    # ── Windows 로컬 절대 경로 (개발 편의) ──
    # 2026판 (현행)
    r"C:\Users\Jini\Desktop\1. 영세사업장 자율점검\3. 취업규칙\취업규칙 마스터 db (2026).xlsx",
    r"E:\취업규칙 마스터 db (2026).xlsx",
    # 2025판 (구판 백업)
    r"C:\Users\Jini\Desktop\1. 영세사업장 자율점검\3. 취업규칙\취업규칙 마스터 db.xlsx",
    r"E:\취업규칙 마스터 db.xlsx",
    r"C:\Users\Jini\Desktop\1. 영세사업장 자율점검\취업규칙 마스터 db.xlsx",
    r"C:\Users\Jini\Desktop\1. 영세사업장 자율점검\취업규칙\취업규칙 마스터 db.xlsx",
]


def _resolve_path(explicit: str | Path | None = None) -> Path:
    if explicit:
        return Path(explicit)
    env = os.environ.get("MASTER_DB_PATH")
    if env:
        return Path(env)

    # 관리자 설정 우선 — master_db_version: "2025" | "2026"
    try:
        from cgr.web.admin.store.settings_store import get as _admin_get
        ver = _admin_get("master_db_version")
        if ver:
            for c in DEFAULT_PATHS:
                if str(ver) in c and Path(c).exists():
                    return Path(c)
    except Exception:
        pass

    for c in DEFAULT_PATHS:
        if Path(c).exists():
            return Path(c)
    raise FileNotFoundError(
        "마스터 DB 위치를 찾지 못함. 다음 중 하나로 지정:\n"
        "  - 인자 path\n"
        "  - MASTER_DB_PATH 환경변수\n"
        f"  - 기본 경로 중 하나: {DEFAULT_PATHS}"
    )


def _load_via_openpyxl(path: Path) -> dict[int, dict[str, Any]] | None:
    """openpyxl 로 시도. 스타일 오류 등 실패 시 None."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path))
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
        out: dict[int, dict[str, Any]] = {}
        for r in range(2, ws.max_row + 1):
            num = ws.cell(r, COLS["no"]).value
            if not isinstance(num, int):
                continue
            row: dict[str, Any] = {}
            for name, col_idx in COLS.items():
                if name == "no":
                    continue
                row[name] = ws.cell(r, col_idx).value
            out[num] = row
        return out
    except Exception:
        return None


def _load_via_zip(path: Path) -> dict[int, dict[str, Any]]:
    """openpyxl 실패 시 raw zip + sharedStrings 로 직접 파싱 (빈출지적 xlsx에서 검증된 방식)."""
    import re
    import xml.etree.ElementTree as ET
    import zipfile

    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(str(path)) as z:
        with z.open("xl/sharedStrings.xml") as f:
            sst = ET.parse(f).getroot()
        ss = []
        for si in sst.findall(f"{ns}si"):
            ss.append("".join(t.text or "" for t in si.iter(f"{ns}t")))
        # find first sheet
        sheet_path = "xl/worksheets/sheet1.xml"
        with z.open(sheet_path) as f:
            sh = ET.parse(f).getroot()

    by_col_letter = {
        "A": "no", "B": "title", "C": "scope", "D": "body",
        "E": "guide", "F": "note", "G": "law", "H": "topic",
        "I": "penalty", "J": "amend_slot", "K": "amend_new", "L": "amend_old",
        "M": "freq_clause", "N": "freq_issue",
    }
    out: dict[int, dict[str, Any]] = {}
    for row in sh.iter(f"{ns}row"):
        ridx = int(row.get("r"))
        if ridx < 2:
            continue
        rec: dict[str, Any] = {}
        no_val = None
        for c in row.findall(f"{ns}c"):
            ref = c.get("r")
            t = c.get("t")
            v_el = c.find(f"{ns}v")
            is_el = c.find(f"{ns}is")
            if t == "s" and v_el is not None:
                val = ss[int(v_el.text)]
            elif t == "inlineStr" and is_el is not None:
                val = "".join(t2.text or "" for t2 in is_el.iter(f"{ns}t"))
            elif v_el is not None:
                val = v_el.text
            else:
                val = None
            m = re.match(r"([A-Z]+)", ref)
            col_letter = m.group(1) if m else ""
            field = by_col_letter.get(col_letter)
            if not field:
                continue
            if field == "no":
                try:
                    no_val = int(val) if val is not None else None
                except (TypeError, ValueError):
                    no_val = None
            else:
                rec[field] = val
        if isinstance(no_val, int):
            out[no_val] = rec
    return out


class MasterDB:
    def __init__(self, path: str | Path | None = None) -> None:
        self.path = _resolve_path(path)
        loaded = _load_via_openpyxl(self.path)
        if loaded is None:
            loaded = _load_via_zip(self.path)
        self._rows = loaded

    # ─── 조회 ─────────────────────────────────────────────
    def article(self, num: int) -> dict[str, Any] | None:
        return self._rows.get(num)

    def title(self, num: int) -> str:
        a = self.article(num)
        if a and a.get("title"):
            return str(a["title"]).strip()
        return f"제{num}조"

    def is_required(self, num: int) -> bool:
        a = self.article(num)
        if not a or not a.get("scope"):
            return False
        return "필수" in str(a["scope"])

    def article_titles(self) -> dict[int, str]:
        return {n: self.title(n) for n in self._rows}

    def all_articles(self) -> list[int]:
        return sorted(self._rows.keys())

    # ─── 셀 단위 헬퍼 ─────────────────────────────────────
    def _cell(self, num: int, field: str) -> str:
        a = self.article(num) or {}
        v = a.get(field)
        return str(v).strip() if v else ""

    def body(self, num: int) -> str:        return self._cell(num, "body")
    def note(self, num: int) -> str:        return self._cell(num, "note")
    def law(self, num: int) -> str:         return self._cell(num, "law")
    def topic(self, num: int) -> str:       return self._cell(num, "topic")
    def penalty(self, num: int) -> str:     return self._cell(num, "penalty")
    def amend_new(self, num: int) -> str:   return self._cell(num, "amend_new")
    def amend_old(self, num: int) -> str:   return self._cell(num, "amend_old")
    def freq_issue(self, num: int) -> str:  return self._cell(num, "freq_issue")
    def freq_clause(self, num: int) -> str: return self._cell(num, "freq_clause")


@lru_cache(maxsize=1)
def get_master_db(path: str | None = None) -> MasterDB:
    """프로세스 단위 단일 인스턴스 (lru_cache로 1회만 로드)."""
    return MasterDB(path)
