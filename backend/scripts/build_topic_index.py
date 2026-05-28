# -*- coding: utf-8 -*-
"""AI노동법 상담 지식데이터 xlsx → 단일 JSON 인덱스 (하드코딩용).

각 시트(31개 토픽)의 content_num + content 를 파싱해서 section_no 추출.
section_no 는 content 시작 부분의 'N.M.K' 패턴에서 캡처.

출력: data/topic_index.json
구조:
  {
    "토픽이름 section_no": {
      "title": "...",
      "content": "..."
    },
    ...
  }
"""
import json
import re
import sys
import io
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

from openpyxl import load_workbook

XLSX_PATH = r"E:/AI노동법 상담 지식데이터_260428.xlsx"
OUT_PATH = Path(r"C:/Users/Jini/Desktop/1. 영세사업장 자율점검/3. 취업규칙/mvp/data/topic_index.json")

# 시트명 → 토픽 키 (slot 의 topic_meta 와 매칭되는 이름)
SHEET_TO_TOPIC = {
    "가산수당": "가산수당",
    "간주근로제": "간주근로제",
    "감시단속적근로": "감시단속적근로",
    "계속근로기간": "계속근로기간",
    "교대제": "교대제",
    "근로시간": "근로시간",
    "근로자성": "근로자성",
    "대지급금": "대지급금",
    "법정외수당": "법정외수당",
    "보상휴가": "보상휴가",
    "상시근로자수": "상시근로자수",
    "선택근로제": "선택근로제",
    "실업급여": "실업급여",
    "연차수당": "연차수당",
    "연차유급휴가": "연차유급휴가",
    "연차촉진": "연차촉진",
    "일용직": "일용직",
    "임금": "임금",
    "임금대장-임금명세서": "임금대장-임금명세서",
    "임금체불": "임금체불",
    "재량근로제": "재량근로제",
    "주휴수당": "주휴수당",
    "최저임금": "최저임금",
    "탄력근로제": "탄력근로제",
    "통상임금": "통상임금",
    "퇴직금": "퇴직금",
    "평균임금": "평균임금",
    "포괄임금제": "포괄임금제",
    "해고예고수당": "해고예고수당",
    "휴업수당": "휴업수당",
    "휴일-휴일대체": "휴일-휴일대체",
}

# section_no 추출 regex — 본문 첫 줄 "1.", "1.1", "2.1.1" 등
_SEC_PATTERN = re.compile(r"^\s*(\d+(?:\.\d+)+)[\.\s]")
_TOP_PATTERN = re.compile(r"^\s*(\d+)\.\s")  # "1. 문서의 목적" 같은 1차 헤더


def parse_content(text: str) -> tuple[str | None, str, str]:
    """content 텍스트 → (section_no, title, body).

    title 은 첫 줄, body 는 그 이후. section_no 가 없으면 None.
    """
    if not text:
        return None, "", ""
    text = str(text).strip()
    lines = text.split("\n", 1)
    first = lines[0].strip()
    rest = lines[1].strip() if len(lines) > 1 else ""

    # 2.1.1 같은 다단계 — 첫 토큰이 section_no
    m = _SEC_PATTERN.match(first)
    if m:
        sec = m.group(1)
        title = first[m.end():].strip()
        return sec, title or "", rest
    # 1. 같은 1차 헤더
    m2 = _TOP_PATTERN.match(first)
    if m2:
        sec = m2.group(1)
        title = first[m2.end():].strip()
        return sec, title or "", rest
    return None, first, rest


def main():
    wb = load_workbook(XLSX_PATH, data_only=True)
    index: dict[str, dict[str, str]] = {}
    stats = {"total_rows": 0, "indexed": 0, "skipped": 0}

    for sheet_name, topic_name in SHEET_TO_TOPIC.items():
        if sheet_name not in wb.sheetnames:
            print(f"!! 시트 없음: {sheet_name}")
            continue
        ws = wb[sheet_name]
        headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        col_content = headers.get("content")
        if not col_content:
            continue
        for r in range(2, ws.max_row + 1):
            stats["total_rows"] += 1
            text = ws.cell(r, col_content).value
            if not text:
                stats["skipped"] += 1
                continue
            sec_no, title, body = parse_content(text)
            if not sec_no:
                stats["skipped"] += 1
                continue
            key = f"{topic_name} {sec_no}"
            # 중복이면 첫 항목 유지 (보통 첫 게 가장 핵심 정의)
            if key in index:
                continue
            index[key] = {
                "topic": topic_name,
                "section_no": sec_no,
                "title": title,
                "content": body or title,
            }
            stats["indexed"] += 1

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(
        json.dumps(index, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"\n총 행: {stats['total_rows']}")
    print(f"인덱스화: {stats['indexed']}")
    print(f"스킵: {stats['skipped']}")
    print(f"\n저장: {OUT_PATH}")
    print(f"\n샘플:")
    for key in list(index.keys())[:5]:
        s = index[key]
        print(f"  '{key}': {s['title'][:50]}")
    # 슬롯에서 자주 쓰는 키 검증
    test_keys = [
        "연차유급휴가 2.1.2",
        "연차촉진 2.1.1",
        "임금 3.1.1",
        "가산수당 2.2.1",
        "임금체불 3.1.1",
        "근로시간 2.1.1",
    ]
    print(f"\n슬롯 매핑 검증:")
    for k in test_keys:
        v = index.get(k)
        if v:
            print(f"  ✓ '{k}': {v['title'][:50]}")
        else:
            print(f"  ✗ '{k}': 미발견")


if __name__ == "__main__":
    main()
