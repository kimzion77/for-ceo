"""(A) 단계 후속 — 2026 본문 갱신 7건을 빈출지적 영역(M·N)에도 반영.

논리:
  대부분의 사업장 취업규칙은 2025 기준으로 작성되어 있으므로,
  2026 변경 사항(노동절·피해사원등·특별교육 자구 정비 등)은 인스펙터가 점검 시
  새로운 빈출 지적 포인트가 된다.

작업:
  1. M(관련 규정 빈출지적): 본문 인용에 들어 있는 2025 표현을 2026 표현으로 치환
     — 단순 결정적 string replace (수동 매핑 테이블)
  2. N(빈출 지적사항): LLM 으로 신규 점검 포인트 생성 후 기존 N 끝에 구분선 + append

대상: 제32, 62, 69, 76, 80, 81, 89조
산출물: 같은 '취업규칙 마스터 db (2026).xlsx' 위에 추가 갱신 + mvp/output/2026_inspect_apply.md
"""
from __future__ import annotations

import json
import sys
import time
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
MVP_ROOT = SCRIPT_DIR.parent
if str(MVP_ROOT) not in sys.path:
    sys.path.insert(0, str(MVP_ROOT))

from cgr import llm_cache
from cgr.config import get_api_key, get_llm_model
from cgr.master_db import COLS, get_master_db
from cgr.parsers.hwp import parse_hwp

from mvp.scripts.compare_body_2026 import strict_body_from_hwp_block
from mvp.scripts.refresh_law_penalty_2026 import (
    DEFAULT_HWP,
    DEFAULT_MASTER,
    DEFAULT_OUT_XLSX,
    extract_body_region,
    split_into_articles,
)


# M(인용 본문) 안의 2025 → 2026 단순 자구 치환 테이블
# 각 항목: (조번호, [(2025_원문, 2026_갱신문), ...])
M_REPLACEMENTS: dict[int, list[tuple[str, str]]] = {
    32: [("근로자의 날(5월 1일)", "노동절(5월 1일)")],
    80: [("피해자가 요청하면", "피해사원등이 요청하면")],
    81: [("피해근로자의 의견", "피해사원등의 의견")],
    89: [
        ("채용 시의 교육", "채용 시 교육"),
        ("작업내용 변경 시의 교육", "작업내용 변경 시 교육"),
        ("유해위험 작업에 사용 시 특별안전 교육", "유해·위험 작업에 필요한 특별교육"),
    ],
}


# N(빈출지적) 신규 포인트 생성 — LLM 시스템 프롬프트
_N_SYSTEM_PROMPT = """당신은 한국 노동법 사업장 취업규칙 검수 전문가다.

[배경]
- 대부분의 사업장 취업규칙은 2025년 표준에 맞춰 작성되어 있다.
- 2026년 표준취업규칙에서 자구·용어가 갱신되었으므로, 사업장 취업규칙이 여전히
  2025 표현 그대로면 검토자가 지적해야 할 새로운 '빈출 지적사항' 이 된다.

[작성 원칙]
- 인스펙터(검토자)가 실제 점검 시 사용할 수 있는 instruction 형식으로 작성.
- 한 두 문장(80~200자). 객관적 사실만 기재, 평가·추측 금지.
- 형식 예시:
  "사업장 취업규칙에 '근로자의 날'로 표기된 경우 2026 개정에 맞춰 '노동절'로 수정 지시."
  "재심회의 개최 후 결과를 신청자에게 통보하는 의무 조항이 누락된 경우 추가 지시."
- 한 줄로 끝내고, 번호(① ② 등)는 붙이지 않는다 — append 시 호스트가 알아서 처리.

[결정성]
- temperature=0. 같은 입력 → 같은 출력.
"""


def _build_n_schema() -> dict:
    return {
        "type": "object",
        "additionalProperties": False,
        "required": ["inspect_point"],
        "properties": {
            "inspect_point": {
                "type": "string",
                "description": "사업장 취업규칙 검토 시 적용할 신규 빈출 지적 포인트 (한 두 문장)",
            }
        },
    }


def _llm_inspect_point(num: int, title: str, summary: str, d_2025: str, d_2026: str) -> str:
    model = get_llm_model()
    schema = _build_n_schema()
    user = (
        f"=== 조 ===\n제{num}조 ({title})\n\n"
        f"=== 2025 → 2026 변경 요지 ===\n{summary}\n\n"
        f"=== 2025 본문 ===\n{d_2025}\n\n"
        f"=== 2026 본문 ===\n{d_2026}\n\n"
        f"위 변경에 따라 사업장 취업규칙이 여전히 2025 표현이거나 신규 의무가 누락된 경우, "
        f"인스펙터가 적용할 수 있는 '신규 빈출 지적사항' 한 항목을 inspect_point 인자로 제출하라."
    )
    cache_key = llm_cache.make_key(_N_SYSTEM_PROMPT, user, schema, model)
    cached = llm_cache.get(cache_key)
    if cached is not None:
        return cached.get("inspect_point", "")
    from openai import APIConnectionError, APITimeoutError, OpenAI, RateLimitError

    client = OpenAI(api_key=get_api_key(), timeout=60.0)
    tools = [{
        "type": "function",
        "function": {"name": "submit", "description": "신규 빈출지적 포인트 제출", "parameters": schema},
    }]
    backoff = (2.0, 5.0, 10.0)
    for attempt in range(3):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": _N_SYSTEM_PROMPT},
                    {"role": "user", "content": user},
                ],
                tools=tools,
                tool_choice={"type": "function", "function": {"name": "submit"}},
                temperature=0,
                top_p=1,
            )
            msg = resp.choices[0].message
            if not msg.tool_calls:
                raise RuntimeError("tool_call 없음")
            payload = json.loads(msg.tool_calls[0].function.arguments)
            llm_cache.put(cache_key, payload)
            return payload.get("inspect_point", "")
        except (APITimeoutError, APIConnectionError, RateLimitError):
            if attempt < 2:
                time.sleep(backoff[attempt])
                continue
            raise
    return ""


def _append_n(existing: str | None, new_point: str) -> str:
    """기존 N에 신규 빈출지적 포인트를 append. 같은 마커 블록이 있으면 교체."""
    new_point = new_point.strip()
    if not new_point:
        return (existing or "").strip()
    sep = "— 2026 갱신 점검 —"
    import re
    if existing and existing.strip():
        cleaned = re.sub(rf"\n*{re.escape(sep)}.*\Z", "", existing, flags=re.DOTALL).rstrip()
        return f"{cleaned}\n\n{sep}\n{new_point}".strip()
    return f"{sep}\n{new_point}"


# 7개 조의 변경 요약 (apply_body_2026 와 동일 — LLM 캐시 통해 동일 결과)
def _summary_for(num: int, title: str, d_2025: str, d_2026: str) -> str:
    """apply_body_2026 의 _llm_summarize 와 동일 — 캐시 hit 만 사용."""
    from mvp.scripts.apply_body_2026 import _llm_summarize
    return _llm_summarize(num, title, d_2025, d_2026)


def main() -> int:
    target_articles = [32, 62, 69, 76, 80, 81, 89]
    md_out = MVP_ROOT / "output" / "2026_inspect_apply.md"

    print(f"master 로드: {DEFAULT_MASTER}")
    db = get_master_db(str(DEFAULT_MASTER))

    print(f"HWP 파싱: {DEFAULT_HWP}")
    arts2026 = split_into_articles(extract_body_region(parse_hwp(DEFAULT_HWP)))

    out_xlsx = Path(DEFAULT_OUT_XLSX)
    if not out_xlsx.exists():
        print(f"[오류] 출력 xlsx 없음: {out_xlsx}")
        return 2

    from openpyxl import load_workbook
    from openpyxl.comments import Comment

    wb = load_workbook(str(out_xlsx))
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    row_by_no: dict[int, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, COLS["no"]).value
        try:
            row_by_no[int(v)] = r
        except (TypeError, ValueError):
            continue

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    model = get_llm_model()
    records: list[dict] = []

    for n in target_articles:
        title = db.title(n)
        d_2025 = db.body(n) or ""  # 캐시 키 일관성 위해 — 단, 이미 D는 2026 으로 갱신됨!
        # 따라서 "2025 D" 는 master 원본에서 다시 읽어야 함
        original_db = get_master_db(str(DEFAULT_MASTER))
        d_2025 = original_db.body(n) or ""
        d_2026 = strict_body_from_hwp_block(arts2026[n])

        # 1. M 인용 본문 자구 치환
        m_old = ws.cell(row_by_no[n], COLS["freq_clause"]).value or ""
        m_new = m_old
        replaced_pairs: list[tuple[str, str]] = []
        for src, dst in M_REPLACEMENTS.get(n, []):
            if src in m_new:
                m_new = m_new.replace(src, dst)
                replaced_pairs.append((src, dst))

        # 2. N 신규 빈출지적 추가
        summary = _summary_for(n, title, d_2025, d_2026)
        inspect_point = _llm_inspect_point(n, title, summary, d_2025, d_2026)
        n_old = ws.cell(row_by_no[n], COLS["freq_issue"]).value or ""
        n_new = _append_n(n_old, inspect_point)

        # 셀 갱신
        if m_new != m_old:
            m_cell = ws.cell(row_by_no[n], COLS["freq_clause"])
            m_cell.value = m_new
            m_cell.comment = Comment(
                f"[2026 자구 치환] {ts}\n"
                + "\n".join(f"  {a} → {b}" for a, b in replaced_pairs),
                "apply_inspect_2026",
            )
        if inspect_point:
            n_cell = ws.cell(row_by_no[n], COLS["freq_issue"])
            n_cell.value = n_new
            n_cell.comment = Comment(
                f"[2026 갱신 점검 추가] {ts}\nmodel: {model}\n포인트: {inspect_point[:300]}",
                "apply_inspect_2026",
            )

        print(f"  제{n}조 ({title})")
        if replaced_pairs:
            for a, b in replaced_pairs:
                print(f"    M 치환: {a!r} → {b!r}")
        else:
            print(f"    M 치환: 없음")
        print(f"    N 신규: {inspect_point}")

        records.append({
            "no": n,
            "title": title,
            "summary": summary,
            "m_old": m_old,
            "m_new": m_new,
            "m_replaced": replaced_pairs,
            "n_old": n_old,
            "n_new": n_new,
            "inspect_point": inspect_point,
        })

    wb.save(str(out_xlsx))
    print(f"\n저장: {out_xlsx}")

    # 보고서
    md_out.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    lines.append("# 2026 본문 갱신에 따른 빈출지적(M·N) 반영 결과\n")
    lines.append(f"- 생성: {ts}")
    lines.append(f"- 반영 조: {len(records)}개")
    lines.append("\n---\n")
    for r in records:
        lines.append(f"## 제{r['no']}조 ({r['title']})")
        lines.append("")
        lines.append(f"**갱신요지**: {r['summary']}")
        lines.append("")
        lines.append("### M(관련 규정) 자구 치환")
        if r["m_replaced"]:
            for a, b in r["m_replaced"]:
                lines.append(f"- `{a}` → `{b}`")
        else:
            lines.append("- (해당 없음)")
        lines.append("")
        lines.append("### N(빈출 지적사항) 추가 포인트")
        lines.append("```")
        lines.append(r["inspect_point"])
        lines.append("```")
        lines.append("")
        lines.append("**N 최종값**:")
        lines.append("```")
        lines.append(r["n_new"])
        lines.append("```")
        lines.append("\n---\n")
    md_out.write_text("\n".join(lines), encoding="utf-8")
    print(f"보고서: {md_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
