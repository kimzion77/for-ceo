"""노무 가이드 DB 시드 — "영세사업주를 위한 꿀팁.xlsx" → SQLite.

자율점검 본질 (사업주 자가 점검) 에 맞춰 **분쟁·신고·구제 신청 류는 시드 자체에서 SKIP**.
다음 13행은 절대 저장 안 함:
  - 서식: FRM020 임금체불 진정서, FRM021 근기법 위반 진정서,
          FRM022 직장 내 괴롭힘 진정서, FRM023 직장 내 성희롱 진정서,
          FRM024 부당해고 구제신청서
  - 기관: ORG004 중앙노동위, ORG005 지방노동위, ORG015 여성긴급전화,
          ORG016 법률구조공단, ORG017 권익위, ORG018 검찰,
          ORG019 법원 민사, ORG021 변호사

ORG002 지방고용노동청 은 **사업주가 근로감독 받는** 측면 가치 있어 유지.

15 시트를 9 테이블에 매핑:
  1·2·3 막막영역 → guide_item (audience 컬럼으로 통합)
  4 사업주 의무 → obligation_timeline
  5 계산 공식 → wage_calc_formula
  6 용어사전 → guide_glossary
  7 법령 근거 → (기존 law_article 보강, 별도 테이블 없음)
  8 신청서식 → form_template (필터)
  9 관할기관 → gov_org (필터)
  10·12 감독 → audit_guide (kind 컬럼)
  11 비치서류 → required_document
  13 채용절차 → recruit_compliance
  14 규모별 → size_threshold_duty
  15 라이프사이클 → employment_lifecycle
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import openpyxl  # noqa: E402

from cgr import db as _db  # noqa: E402


XLSX_PATH = ROOT / "data" / "영세사업주를 위한 꿀팁.xlsx"

# 자율점검 정신상 제외 — 사업주가 사용할 일이 없거나 분쟁 트리거 가능 항목
EXCLUDED_FORM_CODES = {
    "FRM020", "FRM021", "FRM022", "FRM023", "FRM024",
}
EXCLUDED_ORG_CODES = {
    "ORG004", "ORG005",  # 노동위
    "ORG015",            # 여성긴급전화
    "ORG016",            # 법률구조공단
    "ORG017",            # 권익위
    "ORG018",            # 검찰
    "ORG019",            # 법원 민사
    "ORG021",            # 변호사
}


def _rows(ws):
    """헤더 빼고 데이터 행만. None 셀은 빈 문자열로."""
    header: list[str] | None = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = [str(c) if c is not None else "" for c in row]
            continue
        if not any(c is not None and str(c).strip() for c in row):
            continue
        yield dict(
            zip(
                header or [],
                ["" if c is None else str(c).strip() for c in row],
            )
        )


def _audience_from_code(code: str) -> str:
    """WRK001 → worker, EMP001 → employer, COM001 → both."""
    if code.startswith("WRK"):
        return "worker"
    if code.startswith("EMP"):
        return "employer"
    if code.startswith("COM"):
        return "both"
    return "both"


def seed_guide_items(conn, wb) -> int:
    """1·2·3 시트 통합 → guide_item."""
    n = 0
    sheet_specs = [
        ("1_근로자_막막영역", "근로자가 막막한 이유", "worker"),
        ("2_사업주_막막영역", "사업주가 막막한 이유", "employer"),
        ("3_공통_막막영역", None, "both"),
    ]
    for sheet_name, reason_col, default_audience in sheet_specs:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for r in _rows(ws):
            code = r.get("ID", "").strip()
            if not code:
                continue
            audience = _audience_from_code(code) or default_audience
            worker_reason = ""
            employer_reason = ""
            if sheet_name == "1_근로자_막막영역":
                worker_reason = r.get(reason_col or "", "")
            elif sheet_name == "2_사업주_막막영역":
                employer_reason = r.get(reason_col or "", "")
            else:
                # 3 시트 — 근로자 관점·사업주 관점 컬럼 둘 다 있음
                worker_reason = r.get("근로자 관점", "")
                employer_reason = r.get("사업주 관점", "")
            conn.execute(
                "INSERT INTO guide_item "
                "(code, audience, category, title, worker_reason, employer_reason, "
                " key_points, related_laws, priority, applies_under_5, note) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) "
                "ON CONFLICT(code) DO UPDATE SET "
                "  audience = excluded.audience, "
                "  category = excluded.category, "
                "  title = excluded.title, "
                "  key_points = excluded.key_points",
                (
                    code,
                    audience,
                    r.get("카테고리", ""),
                    r.get("항목명", ""),
                    worker_reason,
                    employer_reason,
                    r.get("핵심 포인트", "") or r.get("핵심 판단 기준", ""),
                    r.get("관련 법령", ""),
                    r.get("우선순위", ""),
                    r.get("5인미만 적용", ""),
                    r.get("비고", "") or r.get("위반 시 제재", ""),
                ),
            )
            n += 1
    return n


def seed_obligations(conn, wb) -> int:
    """4_사업주_법령상_의무."""
    if "4_사업주_법령상_의무" not in wb.sheetnames:
        return 0
    n = 0
    for r in _rows(wb["4_사업주_법령상_의무"]):
        code = r.get("ID", "").strip()
        if not code:
            continue
        conn.execute(
            "INSERT INTO obligation_timeline "
            "(code, stage, duty, description, deadline, legal_basis, priority, penalty) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(code) DO UPDATE SET stage = excluded.stage, duty = excluded.duty",
            (
                code,
                r.get("시기", ""),
                r.get("의무 사항", ""),
                r.get("법령상 내용", ""),
                r.get("이행 시점", ""),
                r.get("근거 법령", ""),
                r.get("우선순위", ""),
                r.get("미이행 시 제재", ""),
            ),
        )
        n += 1
    return n


def seed_wage_calc(conn, wb) -> int:
    """5_통상임금_연관계산 — V003~V006 룰 코드 자동 매핑."""
    if "5_통상임금_연관계산" not in wb.sheetnames:
        return 0
    # 계산명 → violation_code 매핑 (수동)
    NAME_TO_V = {
        "연장근로수당": "V003",
        "야간근로수당": "V004",
        "휴일근로수당": "V005",
        "주휴수당": "V006",
    }
    n = 0
    for r in _rows(wb["5_통상임금_연관계산"]):
        code = r.get("ID", "").strip()
        if not code:
            continue
        calc_name = r.get("계산명", "")
        related_v = None
        for k, v in NAME_TO_V.items():
            if k in calc_name:
                related_v = v
                break
        conn.execute(
            "INSERT INTO wage_calc_formula "
            "(code, category, calc_name, formula, conditions, limits, "
            " legal_basis, note, related_violation_code) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(code) DO UPDATE SET calc_name = excluded.calc_name, "
            "  formula = excluded.formula",
            (
                code,
                r.get("카테고리", ""),
                calc_name,
                r.get("계산 공식 (간이)", ""),
                r.get("지급 조건", ""),
                r.get("상한·하한", ""),
                r.get("관련 법령", ""),
                r.get("비고", ""),
                related_v,
            ),
        )
        n += 1
    return n


def seed_glossary(conn, wb) -> int:
    """6_용어사전."""
    if "6_용어사전" not in wb.sheetnames:
        return 0
    n = 0
    for r in _rows(wb["6_용어사전"]):
        code = r.get("ID", "").strip()
        term = r.get("용어", "").strip()
        if not code or not term:
            continue
        conn.execute(
            "INSERT INTO guide_glossary "
            "(code, term, short_def, full_def, confusable_with, legal_basis) "
            "VALUES (?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(code) DO UPDATE SET term = excluded.term, "
            "  short_def = excluded.short_def",
            (
                code,
                term,
                r.get("간이 정의", ""),
                r.get("상세 설명", ""),
                r.get("헷갈리는 용어와의 차이", ""),
                r.get("관련 법령", ""),
            ),
        )
        n += 1
    return n


def _form_download_url(category: str, form_name: str) -> str:
    """카테고리별 공식 자료실 URL 매핑 (Phase 18).

    원칙:
      1) 외부 정부·공공기관 자료실 URL 만 반환 — 우리 서비스는 양식을 호스팅하지 않음
      2) 직링크가 바뀌어도 안전하도록 **자료실 메인 또는 검색 URL** 사용
      3) 사용자가 클릭하면 해당 부처에서 양식을 직접 찾을 수 있도록

    매핑:
      근로계약·취업규칙·퇴직 → 고용노동부 정책자료실 검색
      4대보험 → 4대사회보험 정보연계센터 자료실
      출산·육아·실업급여 → 고용보험 (ei.go.kr)
      산재 → 근로복지공단 (kcomwel.or.kr)
      외국인 근로 → 외국인고용관리시스템 (EPS)
    """
    from urllib.parse import quote
    cat = category or ""
    q = quote(form_name) if form_name else ""

    # 1) 고용노동부 정책자료실 검색 — 표준근로계약서·취업규칙·퇴직 관련
    if any(k in cat for k in ("근로계약", "취업규칙", "퇴직")):
        # 정책자료실 게시판 + 제목 검색
        return (
            "https://www.moel.go.kr/policy/policydata/list.do"
            f"?searchType=tit&searchWord={q}"
        )

    # 2) 4대사회보험 정보연계센터
    if "4대보험" in cat:
        # 두루누리 지원은 별도 사이트
        if "두루누리" in form_name:
            return "https://insurancesupport.or.kr"
        return "https://www.4insure.or.kr/ins4/ptl/data/inseDataList.do"

    # 3) 출산·육아·실업급여 — 고용보험
    if any(k in cat for k in ("출산", "육아", "실업급여")):
        return "https://www.ei.go.kr/ei/eih/cm/hm/main.do"

    # 4) 산재 — 근로복지공단
    if "산재" in cat:
        return "https://www.kcomwel.or.kr/comwel/cust/dataroom/dataRoom.jsp"

    # 5) 외국인 근로 — EPS
    if "외국인" in cat:
        return "https://www.eps.go.kr"

    # 폴백 — 고용노동부 통합검색
    return f"https://www.moel.go.kr/search/totalSearch.do?kwd={q}"


def seed_forms(conn, wb) -> int:
    """8_신청서식 — 분쟁 진정서·구제신청서는 SKIP.

    Phase 18: 카테고리별 공식 자료실 URL 을 `download_url` 에 채움.
    우리 서비스는 양식 호스팅 없음 — 외부 정부 사이트로 안내만.
    """
    if "8_신청서식" not in wb.sheetnames:
        return 0
    n = 0
    skipped = 0
    for r in _rows(wb["8_신청서식"]):
        code = r.get("ID", "").strip()
        if not code:
            continue
        if code in EXCLUDED_FORM_CODES:
            skipped += 1
            continue
        # 제출자 → audience
        submitter = r.get("신청자", "")
        if "사업주" in submitter and "근로자" in submitter:
            audience = "both"
        elif "사업주" in submitter:
            audience = "employer"
        elif "근로자" in submitter:
            audience = "employee"
        else:
            audience = "employer"  # 기본
        category = r.get("카테고리", "")
        form_name = r.get("서식명", "")
        download_url = _form_download_url(category, form_name)
        conn.execute(
            "INSERT INTO form_template "
            "(code, category, form_name, purpose, submitter, submit_to, "
            " submit_method, deadline, legal_basis, download_url, audience) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(code) DO UPDATE SET "
            "  form_name = excluded.form_name, "
            "  download_url = excluded.download_url",
            (
                code,
                category,
                form_name,
                r.get("용도", ""),
                submitter,
                r.get("제출처", ""),
                r.get("제출 방법", ""),
                r.get("제출 기한", ""),
                r.get("근거 법령", ""),
                download_url,
                audience,
            ),
        )
        n += 1
    if skipped:
        print(f"    ※ 분쟁 진정·구제 서식 {skipped}건 SKIP")
    return n


def seed_orgs(conn, wb) -> int:
    """9_관할기관 — 노동위·검찰·법원·변호사 등 분쟁 기관 SKIP."""
    if "9_관할기관" not in wb.sheetnames:
        return 0
    n = 0
    skipped = 0
    for r in _rows(wb["9_관할기관"]):
        code = r.get("ID", "").strip()
        if not code:
            continue
        if code in EXCLUDED_ORG_CODES:
            skipped += 1
            continue
        conn.execute(
            "INSERT INTO gov_org "
            "(code, org_class, org_name, duties, common_cases, phone, "
            " online_channel, jurisdiction, note) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(code) DO UPDATE SET org_name = excluded.org_name",
            (
                code,
                r.get("기관 분류", ""),
                r.get("기관명", ""),
                r.get("담당 업무", ""),
                r.get("주요 처리 민원", ""),
                r.get("연락 방법", ""),
                r.get("온라인 채널", ""),
                r.get("관할 결정 기준", ""),
                r.get("비고", ""),
            ),
        )
        n += 1
    if skipped:
        print(f"    ※ 노동위·검찰·법원·변호사 등 {skipped}건 SKIP")
    return n


def seed_audit_guide(conn, wb) -> int:
    """10_근로감독_종류 + 12_감독_진행절차 통합."""
    n = 0
    if "10_근로감독_종류" in wb.sheetnames:
        for r in _rows(wb["10_근로감독_종류"]):
            code = r.get("ID", "").strip()
            if not code:
                continue
            conn.execute(
                "INSERT INTO audit_guide "
                "(code, kind, name, step_no, timing, description, period_covered, legal_basis) "
                "VALUES (?, 'type', ?, NULL, NULL, ?, ?, ?) "
                "ON CONFLICT(code) DO UPDATE SET name = excluded.name",
                (
                    code,
                    r.get("감독 종류", ""),
                    r.get("실시 사유", ""),
                    r.get("점검 범위 (기간)", ""),
                    r.get("근거 규정", ""),
                ),
            )
            n += 1
    if "12_감독_진행절차" in wb.sheetnames:
        for r in _rows(wb["12_감독_진행절차"]):
            code = r.get("ID", "").strip()
            if not code:
                continue
            step_str = r.get("단계", "").split(".")[0].strip()
            try:
                step_no = int(step_str)
            except ValueError:
                step_no = None
            conn.execute(
                "INSERT INTO audit_guide "
                "(code, kind, name, step_no, timing, description, period_covered, legal_basis) "
                "VALUES (?, 'procedure', ?, ?, ?, ?, NULL, ?) "
                "ON CONFLICT(code) DO UPDATE SET name = excluded.name",
                (
                    code,
                    r.get("단계", ""),
                    step_no,
                    r.get("시점", ""),
                    r.get("절차 내용", ""),
                    r.get("근거 규정", ""),
                ),
            )
            n += 1
    return n


def seed_required_docs(conn, wb) -> int:
    """11_법령상_비치서류."""
    if "11_법령상_비치서류" not in wb.sheetnames:
        return 0
    n = 0
    for r in _rows(wb["11_법령상_비치서류"]):
        code = r.get("ID", "").strip()
        if not code:
            continue
        conn.execute(
            "INSERT INTO required_document "
            "(code, classification, doc_name, description, prep_time, "
            " retention_period, legal_basis, penalty) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(code) DO UPDATE SET doc_name = excluded.doc_name",
            (
                code,
                r.get("분류", ""),
                r.get("서류명", ""),
                r.get("법령상 내용", ""),
                r.get("작성·비치 시점", ""),
                r.get("법정 보존기간", ""),
                r.get("근거 법령", ""),
                r.get("미작성·미비치 시 제재", ""),
            ),
        )
        n += 1
    return n


def seed_recruit(conn, wb) -> int:
    """13_채용절차_준수사항."""
    if "13_채용절차_준수사항" not in wb.sheetnames:
        return 0
    n = 0
    for r in _rows(wb["13_채용절차_준수사항"]):
        code = r.get("ID", "").strip()
        if not code:
            continue
        conn.execute(
            "INSERT INTO recruit_compliance "
            "(code, stage, duty, description, violation_examples, "
            " penalty, applies_to, legal_basis, checkpoint) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(code) DO UPDATE SET duty = excluded.duty",
            (
                code,
                r.get("단계", ""),
                r.get("의무·금지 사항", ""),
                r.get("구체 내용", ""),
                r.get("위반 사례", ""),
                r.get("제재", ""),
                r.get("적용 대상", ""),
                r.get("근거 조문", ""),
                r.get("체크포인트", ""),
            ),
        )
        n += 1
    return n


def seed_size_duty(conn, wb) -> int:
    """14_규모별_의무."""
    if "14_규모별_의무" not in wb.sheetnames:
        return 0
    n = 0
    for r in _rows(wb["14_규모별_의무"]):
        code = r.get("ID", "").strip()
        if not code:
            continue
        conn.execute(
            "INSERT INTO size_threshold_duty "
            "(code, min_size, duty, description, related_docs, legal_basis, note) "
            "VALUES (?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(code) DO UPDATE SET duty = excluded.duty",
            (
                code,
                r.get("적용 시작 규모", ""),
                r.get("의무·적용 사항", ""),
                r.get("내용", ""),
                r.get("관련 서류", ""),
                r.get("근거 법령", ""),
                r.get("비고", ""),
            ),
        )
        n += 1
    return n


def seed_lifecycle(conn, wb) -> int:
    """15_채용부터_종료까지."""
    if "15_채용부터_종료까지" not in wb.sheetnames:
        return 0
    n = 0
    for r in _rows(wb["15_채용부터_종료까지"]):
        code = r.get("ID", "").strip()
        if not code:
            continue
        conn.execute(
            "INSERT INTO employment_lifecycle "
            "(code, phase, sub_topic, requirement, related_docs, timing, legal_basis, note) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?) "
            "ON CONFLICT(code) DO UPDATE SET requirement = excluded.requirement",
            (
                code,
                r.get("단계", ""),
                r.get("국면", ""),
                r.get("법령상 요구사항", ""),
                r.get("관련 서류·기록", ""),
                r.get("관련 시점", ""),
                r.get("근거 법령", ""),
                r.get("안내 참고사항", ""),
            ),
        )
        n += 1
    return n


# ════════════════════════════════════════════════════════
# main
# ════════════════════════════════════════════════════════
def run() -> None:
    if not XLSX_PATH.exists():
        print(f"  (skip) {XLSX_PATH.name} not found")
        return
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    _db.init_schema(drop_first=False)
    with _db.connect() as conn:
        n1 = seed_guide_items(conn, wb)
        print(f"  guide_item: {n1}")
        n2 = seed_obligations(conn, wb)
        print(f"  obligation_timeline: {n2}")
        n3 = seed_wage_calc(conn, wb)
        print(f"  wage_calc_formula: {n3}")
        n4 = seed_glossary(conn, wb)
        print(f"  guide_glossary: {n4}")
        n5 = seed_forms(conn, wb)
        print(f"  form_template: {n5}")
        n6 = seed_orgs(conn, wb)
        print(f"  gov_org: {n6}")
        n7 = seed_audit_guide(conn, wb)
        print(f"  audit_guide: {n7}")
        n8 = seed_required_docs(conn, wb)
        print(f"  required_document: {n8}")
        n9 = seed_recruit(conn, wb)
        print(f"  recruit_compliance: {n9}")
        n10 = seed_size_duty(conn, wb)
        print(f"  size_threshold_duty: {n10}")
        n11 = seed_lifecycle(conn, wb)
        print(f"  employment_lifecycle: {n11}")
        print(
            f"  total: {n1 + n2 + n3 + n4 + n5 + n6 + n7 + n8 + n9 + n10 + n11} rows"
        )


if __name__ == "__main__":
    run()
