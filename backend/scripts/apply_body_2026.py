"""HWP에서 발견된 본문 변경 7건을 마스터 DB의 D/K/L 컬럼에 반영.

대상 (compare_body_2026.py 분석 결과):
  제32조, 제62조, 제69조, 제76조, 제80조, 제81조, 제89조

규칙:
  - D(취업규칙 안)  ← 2026 strict body 로 교체
  - L(구법)         ← 기존 값이 있으면 보존 + '\n\n— 2025 본문 —\n' + 2025 D 부착
  - K(갱신사항)     ← 기존 값이 있으면 보존 + '\n\n— 2026 갱신 —\n' + LLM 요약

이미 G·I가 갱신된 '취업규칙 마스터 db (2026).xlsx' 위에 추가 작업한다.
원본 '취업규칙 마스터 db.xlsx'는 손대지 않는다.

산출물:
  - 같은 '취업규칙 마스터 db (2026).xlsx' (D/K/L 추가 갱신)
  - mvp/output/2026_body_apply.md  (반영 결과 요약)
"""
from __future__ import annotations

import json
import re
import sys
import time
from datetime import datetime
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
MVP_ROOT = SCRIPT_DIR.parent
if str(MVP_ROOT) not in sys.path:
    sys.path.insert(0, str(MVP_ROOT))

from cgr import llm_cache
from cgr.config import get_api_key, get_llm_model
from cgr.master_db import COLS, get_master_db
from cgr.parsers.hwp import parse_hwp

from mvp.scripts.compare_body_2026 import (
    normalize_for_compare,
    strict_body_from_hwp_block,
)
from mvp.scripts.refresh_law_penalty_2026 import (
    DEFAULT_HWP,
    DEFAULT_MASTER,
    DEFAULT_OUT_XLSX,
    extract_body_region,
    split_into_articles,
)


_SYSTEM_PROMPT = """당신은 한국 노동법 전문가다. 주어진 한 조문에 대해, 2025년 표준취업규칙 본문과
2026년 표준취업규칙 본문의 차이를 한 줄(50~120자) 요약으로 정리해 함수 호출로 제출하라.

[작성 원칙]
- 무엇이 어떻게 바뀌었는지 사실만 기재. 추정·평가 금지.
- 형식: "<바뀐 항목>: '<2025>' → '<2026>'" 또는 "<신설 사항 요지>" 식.
- 한국어 정식 용어 사용. 인용은 작은따옴표.
- 결정성 보장 — temperature=0.
"""


def _build_schema() -> dict:
    return {
        "type": "object",
        "additionalProperties": False,
        "required": ["summary"],
        "properties": {"summary": {"type": "string"}},
    }


def _llm_summarize(num: int, title: str, d_2025: str, d_2026: str) -> str:
    model = get_llm_model()
    schema = _build_schema()
    user = (
        f"=== 조 ===\n제{num}조 ({title})\n\n"
        f"=== 2025 본문 ===\n{d_2025}\n\n"
        f"=== 2026 본문 ===\n{d_2026}\n\n"
        f"두 본문의 차이를 1문장 요약으로 submit 함수 인자(summary)에 담아라."
    )
    cache_key = llm_cache.make_key(_SYSTEM_PROMPT, user, schema, model)
    cached = llm_cache.get(cache_key)
    if cached is not None:
        return cached.get("summary", "")
    from openai import APIConnectionError, APITimeoutError, OpenAI, RateLimitError

    client = OpenAI(api_key=get_api_key(), timeout=60.0)
    tools = [{
        "type": "function",
        "function": {"name": "submit", "description": "변경 요약 제출", "parameters": schema},
    }]
    backoff = (2.0, 5.0, 10.0)
    for attempt in range(3):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": _SYSTEM_PROMPT},
                    {"role": "user", "content": user},
                ],
                tools=tools,
                tool_choice={"type": "function", "function": {"name": "submit"}},
                temperature=0,
                top_p=1,
            )
            msg = resp.choices[0].message
            if not msg.tool_calls:
                raise RuntimeError("tool_call 없음")
            payload = json.loads(msg.tool_calls[0].function.arguments)
            llm_cache.put(cache_key, payload)
            return payload.get("summary", "")
        except (APITimeoutError, APIConnectionError, RateLimitError):
            if attempt < 2:
                time.sleep(backoff[attempt])
                continue
            raise
    return ""


def _append_new_with_separator(existing: str | None, new_block: str, label: str) -> str:
    """기존 셀 값에 신규 블록을 구분선으로 덧붙임."""
    new_block = new_block.strip()
    if not new_block:
        return (existing or "").strip()
    sep = f"— {label} —"
    if existing and existing.strip():
        # 이미 같은 label 블록이 있으면 그 부분만 교체
        marker_pat = re.compile(rf"\n*{re.escape(sep)}.*\Z", re.DOTALL)
        cleaned = marker_pat.sub("", existing).rstrip()
        return f"{cleaned}\n\n{sep}\n{new_block}".strip()
    return f"{sep}\n{new_block}"


def main() -> int:
    target_articles = [32, 62, 69, 76, 80, 81, 89]
    md_out = MVP_ROOT / "output" / "2026_body_apply.md"

    print(f"master 로드: {DEFAULT_MASTER}")
    db = get_master_db(str(DEFAULT_MASTER))

    print(f"HWP 파싱: {DEFAULT_HWP}")
    arts2026 = split_into_articles(extract_body_region(parse_hwp(DEFAULT_HWP)))

    out_xlsx = Path(DEFAULT_OUT_XLSX)
    if not out_xlsx.exists():
        print(f"[오류] 출력 xlsx 없음 — 먼저 refresh_law_penalty_2026 실행 필요: {out_xlsx}")
        return 2

    from openpyxl import load_workbook
    from openpyxl.comments import Comment

    wb = load_workbook(str(out_xlsx))
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    # row 매핑
    row_by_no: dict[int, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, COLS["no"]).value
        try:
            row_by_no[int(v)] = r
        except (TypeError, ValueError):
            continue

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    model = get_llm_model()
    records: list[dict] = []

    for n in target_articles:
        title = db.title(n)
        d_2025 = db.body(n) or ""
        d_2026 = strict_body_from_hwp_block(arts2026[n])
        if normalize_for_compare(d_2025) == normalize_for_compare(d_2026):
            print(f"  제{n}조 — 본문 차이 없음, 스킵")
            continue

        # LLM 으로 갱신요약
        summary = _llm_summarize(n, title, d_2025, d_2026)
        print(f"  제{n}조 ({title}) — 요약: {summary}")

        old_K = ws.cell(row_by_no[n], COLS["amend_new"]).value or ""
        old_L = ws.cell(row_by_no[n], COLS["amend_old"]).value or ""

        new_K = _append_new_with_separator(old_K, summary, "2026 갱신")
        new_L = _append_new_with_separator(old_L, d_2025, "2025 본문")

        # 셀 갱신
        d_cell = ws.cell(row_by_no[n], COLS["body"])
        k_cell = ws.cell(row_by_no[n], COLS["amend_new"])
        l_cell = ws.cell(row_by_no[n], COLS["amend_old"])
        d_cell.value = d_2026
        k_cell.value = new_K
        l_cell.value = new_L

        cmt = (
            f"[2026 본문 갱신] {ts}\n"
            f"model: {model}\n"
            f"요약: {summary[:200]}"
        )
        d_cell.comment = Comment(cmt[:1000], "apply_body_2026")
        k_cell.comment = Comment(cmt[:1000], "apply_body_2026")
        l_cell.comment = Comment(cmt[:1000], "apply_body_2026")

        records.append({
            "no": n,
            "title": title,
            "d_2025": d_2025,
            "d_2026": d_2026,
            "summary": summary,
            "old_K": old_K,
            "new_K": new_K,
            "old_L": old_L,
            "new_L": new_L,
        })

    wb.save(str(out_xlsx))
    print(f"\n저장: {out_xlsx}")

    # 보고서 작성
    md_out.parent.mkdir(parents=True, exist_ok=True)
    lines: list[str] = []
    lines.append("# 2026 표준취업규칙 본문(D)·갱신사항(K)·구법(L) 반영 결과\n")
    lines.append(f"- 생성: {ts}")
    lines.append(f"- 반영 조: **{len(records)}개**")
    lines.append(
        "\n> 본 작업은 HWP에서 발견된 본문 차이만 반영함. "
        "HWP 외 입법 변경(예: 난임치료휴가 4일 유급 등)은 추후 K 컬럼에 별도 추가 필요."
    )
    lines.append("\n---\n")
    for r in records:
        lines.append(f"## 제{r['no']}조 ({r['title']})")
        lines.append("")
        lines.append(f"**갱신요약:** {r['summary']}")
        lines.append("")
        lines.append("### D 갱신")
        lines.append("- 2025:")
        lines.append("```")
        lines.append(r["d_2025"])
        lines.append("```")
        lines.append("- 2026:")
        lines.append("```")
        lines.append(r["d_2026"])
        lines.append("```")
        if r["old_K"]:
            lines.append("\n**K(갱신사항) 기존값 보존**: \n```\n" + r["old_K"] + "\n```")
        lines.append("\n**K(갱신사항) 최종값**:\n```\n" + r["new_K"] + "\n```")
        if r["old_L"]:
            lines.append("\n**L(구법) 기존값 보존**:\n```\n" + r["old_L"] + "\n```")
        lines.append("\n---\n")
    md_out.write_text("\n".join(lines), encoding="utf-8")
    print(f"보고서: {md_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
