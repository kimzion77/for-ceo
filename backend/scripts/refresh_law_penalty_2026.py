"""2026년 표준취업규칙 HWP 기반으로 마스터 DB의 관련법령(G)·벌칙조항(I)만 갱신.

산출물:
  - "취업규칙 마스터 db (2026).xlsx" : 2025 원본 복사 후 G·I열만 덮어쓰기
  - mvp/output/2026_law_penalty_diff.md : 98조 변경 전·후 검토표
  - mvp/output/2026_law_penalty_raw.json : LLM 응답 원문 (재실행 캐시)

본 스크립트는 본문(D)·메타(E·F·H·M·N)·갱신(K·L) 등은 일절 손대지 않는다.

사용:
  python -m mvp.scripts.refresh_law_penalty_2026
  python -m mvp.scripts.refresh_law_penalty_2026 --only 33,56,61
  python -m mvp.scripts.refresh_law_penalty_2026 --dry-run     # 캐시 hit 만, 새 LLM 호출 없음
"""
from __future__ import annotations

import argparse
import io
import json
import re
import shutil
import sys
import time
from datetime import datetime
from pathlib import Path

# Windows 콘솔(cp949) 에서 한글 미들닷(․) 등 인코딩 실패 방지 — UTF-8 강제
if sys.platform == "win32":
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

# 프로젝트 루트(mvp/) 를 import path 에 추가 — scripts/ 가 패키지가 아니어도 import 가능
SCRIPT_DIR = Path(__file__).resolve().parent
MVP_ROOT = SCRIPT_DIR.parent
if str(MVP_ROOT) not in sys.path:
    sys.path.insert(0, str(MVP_ROOT))

from cgr import llm_cache
from cgr.config import get_api_key, get_llm_model
from cgr.master_db import COLS, get_master_db
from cgr.parsers.hwp import parse_hwp


# 기본 경로
PROJECT_DIR = MVP_ROOT.parent
DEFAULT_MASTER = PROJECT_DIR / "취업규칙 마스터 db.xlsx"
DEFAULT_HWP = PROJECT_DIR / "표준취업규칙(2026년, 배포).hwp"
DEFAULT_OUT_XLSX = PROJECT_DIR / "취업규칙 마스터 db (2026).xlsx"
DEFAULT_OUT_MD = MVP_ROOT / "output" / "2026_law_penalty_diff.md"
DEFAULT_OUT_JSON = MVP_ROOT / "output" / "2026_law_penalty_raw.json"


# ─── HWP → 조문 분할 ────────────────────────────────────────────

_HEADING = re.compile(r"^제\s*(\d+)\s*조\s*\(.*?\)")
_BUCHIK = re.compile(r"^부\s+칙\s*$")


def extract_body_region(text: str) -> list[str]:
    """HWP 평문에서 본문(제1조~제98조) 영역만 잘라낸다.

    경계:
      시작: '(작성시 착안사항)' 라인 (TOC 끝 직후)
      끝:   첫 '부   칙' 라인
    """
    lines = text.split("\n")
    starts = [i for i, ln in enumerate(lines) if ln.strip() == "(작성시 착안사항)"]
    if not starts:
        raise RuntimeError("HWP에서 본문 시작 앵커 '(작성시 착안사항)' 를 찾지 못함")
    start = starts[0]
    end = next(
        (i for i in range(start + 1, len(lines)) if _BUCHIK.match(lines[i].strip())),
        len(lines),
    )
    return lines[start:end]


def split_into_articles(body: list[str]) -> dict[int, str]:
    """본문 라인을 조 번호별 텍스트 블록으로 분할.

    - 같은 번호가 두 번 나오는 조(제61·62조 대안 조항)는 첫 등장 ~ 다음 번호 등장 직전까지
      모두 한 블록으로 묶는다.
    """
    first_hit: dict[int, int] = {}
    for i, ln in enumerate(body):
        m = _HEADING.match(ln.strip())
        if m:
            n = int(m.group(1))
            first_hit.setdefault(n, i)
    missing = [n for n in range(1, 99) if n not in first_hit]
    if missing:
        raise RuntimeError(f"제 N조 누락: {missing}")
    out: dict[int, str] = {}
    for n in range(1, 99):
        s = first_hit[n]
        e = next((first_hit[m] for m in range(n + 1, 99) if m in first_hit), len(body))
        out[n] = "\n".join(body[s:e]).strip()
    return out


# ─── LLM 추출 ───────────────────────────────────────────────────

_SYSTEM_PROMPT = """당신은 한국 노동법 전문가이다. 2026년 고용노동부 배포 '표준취업규칙' 본문 한 조(條)를 보고,
아래 두 항목을 정확히 추출·정규화하여 함수 호출로 제출하라.

[1] law (관련법령)
- 본문·[필수/선택]·☞(참고) 어디에 등장하든 명시적으로 인용된 법령 조문을 모두 수집.
- 정규화 표기: "법률명 제N조[ 제M항][ 제K호]" 한 줄 단위. 시행령/시행규칙은 동일 규칙으로 별도 표기.
- 동일 조문 중복 제거. 본문에 등장하지 않는 조문은 추가 금지(추측 금지).
- 조 자체에 직접 적용되는 근로기준법 제93조 각 호(필수기재사항)에 해당하면, 해당 호도 포함.
- '☞ (참고) … (근로기준법 제xx조)', '(2018.5월 …)' 등 출처 메타데이터 안의 조문 인용도 포함.

[2] penalty (벌칙조항)
- '이 조의 의무가 위반될 경우' 적용되는 벌칙·과태료 조항을 정리. 본문에 명시 없으면 law 기반으로 도출.
- 다음 매핑은 한국 노동법 상 표준이므로, 해당 의무가 본 조의 핵심이면 자동 적용:
  · 취업규칙 필수기재사항(근로기준법 제93조) 미기재 → 근로기준법 제116조 제2항 제2호: 500만원 이하 과태료
  · 임금 미지급(체불) → 근로기준법 제109조 제1항: 3년 이하 징역 또는 3천만원 이하 벌금
  · 근로조건 위반(근로시간/휴게/휴일/연차/모성보호 등 핵심 의무) → 근로기준법 제110조 제1호 등 해당 조항: 2년 이하 징역 또는 2천만원 이하 벌금
  · 부당해고 등 → 근로기준법 제23조 위반: 노동위 구제(직접 형벌 없음)
  · 직장 내 괴롭힘 미조치(근로기준법 제76조의3) → 제116조 제2항 제2호: 500만원 이하 과태료
  · 성희롱 예방교육 미실시(남녀고용평등법 제13조) → 제39조 제3항 제1호: 500만원 이하 과태료
- 형식: "[위반 트리거] 근거조문: 제재 내용". 여러 항목은 별도 entry로 분리.
- 직접 적용 벌칙이 없는 임의·확인적 규정이면 'direct_penalty=false' 로 표기하고 entries=[].

[정확성 원칙]
- 본문에 인용되지 않은 법령을 임의로 끌어오지 말 것 (단, [2] 의 표준 매핑은 의무가 명백히 본 조의 핵심일 때 적용 가능).
- 2025 master 의 기존 G/I 가 함께 제공된다. 2026 본문에 동일 근거가 있다면 같은 항목을 재사용해 일관성 유지.
- 2025 에 있었으나 2026 본문에서 더 이상 인용되지 않는 조문은 제외. 단 2025 의 표준 벌칙 매핑은 유지.
- 출처 인용(source_quote): 2026 본문에서 핵심 근거 한 문장(50~200자) 발췌. 가공 금지.

[결정성]
- 같은 입력 → 같은 출력. temperature=0 환경.
"""


def _build_tool_schema() -> dict:
    return {
        "type": "object",
        "additionalProperties": False,
        "required": ["law", "penalty", "source_quote"],
        "properties": {
            "law": {
                "type": "array",
                "items": {"type": "string"},
                "description": "정규화된 법령 인용 한 줄 단위 배열",
            },
            "penalty": {
                "type": "object",
                "additionalProperties": False,
                "required": ["direct_penalty", "entries"],
                "properties": {
                    "direct_penalty": {"type": "boolean"},
                    "entries": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "additionalProperties": False,
                            "required": ["trigger", "basis", "sanction"],
                            "properties": {
                                "trigger": {"type": "string"},
                                "basis": {"type": "string"},
                                "sanction": {"type": "string"},
                            },
                        },
                    },
                },
            },
            "source_quote": {"type": "string"},
        },
    }


def _build_user_prompt(num: int, title: str, body_2026: str, g_old: str, i_old: str) -> str:
    return (
        f"=== 조 정보 ===\n"
        f"제{num}조 ({title})\n\n"
        f"=== 2026 표준취업규칙 본문 (HWP 추출) ===\n"
        f"{body_2026}\n\n"
        f"=== 2025 master DB 기존 값 (참고용) ===\n"
        f"[관련법령 G]\n{g_old or '(공란)'}\n\n"
        f"[벌칙조항 I]\n{i_old or '(공란)'}\n\n"
        f"=== 작업 ===\n"
        f"위 2026 본문을 기준으로 [1] law, [2] penalty 를 정규화 추출하여 submit 함수로 제출하라.\n"
        f"2025 값은 일관성 참고용일 뿐, 2026 본문에 근거가 없으면 그대로 두지 말 것."
    )


def _serialize_law(law_list: list[str]) -> str:
    seen = []
    for s in law_list:
        s2 = (s or "").strip()
        if s2 and s2 not in seen:
            seen.append(s2)
    return "\n".join(seen)


def _serialize_penalty(penalty_obj: dict) -> str:
    if not penalty_obj.get("direct_penalty"):
        if not penalty_obj.get("entries"):
            return "직접 적용 벌칙 없음(임의·확인적 규정)"
    entries = penalty_obj.get("entries") or []
    rendered: list[str] = []
    for e in entries:
        trig = (e.get("trigger") or "").strip()
        basis = (e.get("basis") or "").strip()
        sanction = (e.get("sanction") or "").strip()
        if not (trig or basis or sanction):
            continue
        rendered.append(f"[{trig}] {basis}: {sanction}".strip())
    return "\n".join(rendered) if rendered else "직접 적용 벌칙 없음(임의·확인적 규정)"


_NO_PENALTY_PREFIX = "직접 적용 벌칙 없음"


def _is_no_penalty(s: str) -> bool:
    """벌칙 셀이 '직접 적용 벌칙 없음(...)' 또는 빈 문자열인지."""
    if not s or not s.strip():
        return True
    return s.strip().startswith(_NO_PENALTY_PREFIX)


def revert_regressions(records: list[dict]) -> list[int]:
    """LLM이 2025의 실질적 벌칙 항목을 '직접 적용 벌칙 없음'으로 축소한 행을 2025 값으로 복원.

    같은 행의 G(관련법령)도 함께 2025 값으로 복원한다 — 벌칙 트리거가 인용한 법령이
    G에서도 함께 누락되는 경우(예: 56조 '금품청산 위반' ↔ 근로기준법 제36조)가 흔하기 때문.
    record 에 `reverted: True` 마커와 사유를 기록한다.
    """
    reverted: list[int] = []
    for r in records:
        if _is_no_penalty(r.get("i_new", "")) and not _is_no_penalty(r.get("i_old", "")):
            r["g_new"] = r["g_old"]
            r["i_new"] = r["i_old"]
            r["g_new_norm"] = _norm(r["g_new"])
            r["i_new_norm"] = _norm(r["i_new"])
            r["reverted"] = True
            r["revert_reason"] = (
                "LLM이 2025의 실질적 벌칙 항목을 '직접 적용 벌칙 없음'으로 축소 → "
                "기존 큐레이션 보존을 위해 G·I를 2025 값으로 복원"
            )
            reverted.append(r["no"])
        else:
            r.setdefault("reverted", False)
    return reverted


def llm_extract(num: int, title: str, body_2026: str, g_old: str, i_old: str, *, dry_run: bool) -> dict:
    """LLM 호출로 G·I 갱신값 추출. 캐시 우선."""
    model = get_llm_model()
    schema = _build_tool_schema()
    user_prompt = _build_user_prompt(num, title, body_2026, g_old, i_old)
    cache_key = llm_cache.make_key(_SYSTEM_PROMPT, user_prompt, schema, model)
    cached = llm_cache.get(cache_key)
    if cached is not None:
        return cached
    if dry_run:
        raise RuntimeError(f"제{num}조 캐시 미존재 — dry-run 종료")
    # OpenAI 호출
    from openai import APIConnectionError, APITimeoutError, OpenAI, RateLimitError

    client = OpenAI(api_key=get_api_key(), timeout=60.0)
    tools = [
        {
            "type": "function",
            "function": {
                "name": "submit",
                "description": "관련법령·벌칙조항 추출 결과 제출",
                "parameters": schema,
            },
        }
    ]
    backoff = (2.0, 5.0, 10.0)
    last_err: Exception | None = None
    for attempt in range(3):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[
                    {"role": "system", "content": _SYSTEM_PROMPT},
                    {"role": "user", "content": user_prompt},
                ],
                tools=tools,
                tool_choice={"type": "function", "function": {"name": "submit"}},
                temperature=0,
                top_p=1,
            )
            msg = resp.choices[0].message
            if not msg.tool_calls:
                raise RuntimeError(f"LLM tool_call 없음: {msg}")
            payload = json.loads(msg.tool_calls[0].function.arguments)
            llm_cache.put(cache_key, payload)
            return payload
        except (APITimeoutError, APIConnectionError, RateLimitError) as e:
            last_err = e
            if attempt < 2:
                time.sleep(backoff[attempt])
                continue
            raise
    raise RuntimeError(f"LLM 호출 실패: {last_err}")


# ─── 출력 ───────────────────────────────────────────────────────


def write_xlsx(master_path: Path, out_path: Path, records: list[dict], *, model_name: str) -> None:
    """2025 원본 복사 → G(law)·I(penalty)만 덮어쓰기 + Comment."""
    from openpyxl import load_workbook
    from openpyxl.comments import Comment

    if out_path.exists():
        out_path.unlink()
    shutil.copyfile(master_path, out_path)

    wb = load_workbook(str(out_path))
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    # 행 번호(no) → 엑셀 row index 매핑
    row_by_no: dict[int, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, COLS["no"]).value
        try:
            n = int(v)
        except (TypeError, ValueError):
            continue
        row_by_no[n] = r

    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    for rec in records:
        n = rec["no"]
        if n not in row_by_no:
            continue
        r = row_by_no[n]
        g_cell = ws.cell(r, COLS["law"])
        i_cell = ws.cell(r, COLS["penalty"])
        g_cell.value = rec["g_new"]
        i_cell.value = rec["i_new"]
        if rec.get("reverted"):
            cmt_text = (
                f"[2026 LLM 보정] {ts}\n"
                f"사유: {rec.get('revert_reason', '')}\n"
                f"→ G·I를 2025 값으로 복원"
            )
        else:
            cmt_text = (
                f"[2026 갱신] {ts}\n"
                f"model: {model_name}\n"
                f"출처: {rec.get('source_quote', '').strip()[:300]}"
            )
        # Comment 글자수 제한 회피 위해 trim
        g_cell.comment = Comment(cmt_text[:1000], "refresh_law_penalty_2026")
        i_cell.comment = Comment(cmt_text[:1000], "refresh_law_penalty_2026")

    wb.save(str(out_path))


def write_diff_md(out_path: Path, records: list[dict]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    g_changed = sum(1 for r in records if r["g_old_norm"] != r["g_new_norm"])
    i_changed = sum(1 for r in records if r["i_old_norm"] != r["i_new_norm"])
    both = sum(
        1 for r in records
        if r["g_old_norm"] != r["g_new_norm"] and r["i_old_norm"] != r["i_new_norm"]
    )
    reverted_nos = [r["no"] for r in records if r.get("reverted")]
    lines: list[str] = []
    lines.append("# 2026년 표준취업규칙 반영 — 관련법령(G)·벌칙조항(I) 갱신 검토표\n")
    lines.append(f"- 생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append(f"- 대상 조문 수: {len(records)}")
    lines.append(f"- G 변경: **{g_changed}건** / I 변경: **{i_changed}건** / 둘 다: **{both}건**")
    if reverted_nos:
        nos_str = ", ".join(f"제{n}조" for n in reverted_nos)
        lines.append(
            f"- LLM 회귀 보정: **{len(reverted_nos)}건** 2025 값으로 복원 → {nos_str}"
        )
    lines.append("")
    lines.append(
        "> 본 산출물은 G·I 두 컬럼만 갱신함. 본문(D)·착안사항(E)·참고(F)·"
        "연관주제(H)·갱신슬롯(J·K·L)·빈출(M·N) 컬럼은 2025 원본 그대로 유지됨."
    )
    lines.append("")

    for r in records:
        flags = []
        if r["g_old_norm"] != r["g_new_norm"]:
            flags.append("G 변경")
        if r["i_old_norm"] != r["i_new_norm"]:
            flags.append("I 변경")
        if r.get("reverted"):
            flags.append("LLM 회귀 보정(2025 복원)")
        flag_s = f" [{' · '.join(flags)}]" if flags else " (변경 없음)"
        lines.append(f"## 제{r['no']}조 ({r['title']}){flag_s}")
        lines.append("")
        lines.append("**G(관련법령)**")
        lines.append("- 2025:")
        lines.append("```")
        lines.append(r["g_old"] or "(공란)")
        lines.append("```")
        lines.append("- 2026:")
        lines.append("```")
        lines.append(r["g_new"] or "(공란)")
        lines.append("```")
        lines.append("")
        lines.append("**I(벌칙조항)**")
        lines.append("- 2025:")
        lines.append("```")
        lines.append(r["i_old"] or "(공란)")
        lines.append("```")
        lines.append("- 2026:")
        lines.append("```")
        lines.append(r["i_new"] or "(공란)")
        lines.append("```")
        lines.append("")
        sq = (r.get("source_quote") or "").strip()
        if sq:
            lines.append(f"**출처(2026 본문)**: {sq}")
            lines.append("")
        lines.append("---")
        lines.append("")

    out_path.write_text("\n".join(lines), encoding="utf-8")


def write_raw_json(out_path: Path, records: list[dict]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(
        json.dumps(records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ─── 정규화 (비교용) ────────────────────────────────────────────


def _norm(s: str) -> str:
    """공백·줄바꿈만 정규화한 비교 키. 의미 동등성은 비교하지 않음."""
    if not s:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip()


# ─── 메인 ───────────────────────────────────────────────────────


def main() -> int:
    ap = argparse.ArgumentParser(description="2026 HWP 기반 G·I 컬럼 갱신")
    ap.add_argument("--master", default=str(DEFAULT_MASTER))
    ap.add_argument("--hwp", default=str(DEFAULT_HWP))
    ap.add_argument("--out-xlsx", default=str(DEFAULT_OUT_XLSX))
    ap.add_argument("--out-md", default=str(DEFAULT_OUT_MD))
    ap.add_argument("--out-json", default=str(DEFAULT_OUT_JSON))
    ap.add_argument("--only", default="", help="처리할 조 번호 콤마 구분 (예: 33,56,61)")
    ap.add_argument("--dry-run", action="store_true", help="LLM 호출 없이 캐시만 사용")
    args = ap.parse_args()

    master_path = Path(args.master)
    hwp_path = Path(args.hwp)
    out_xlsx = Path(args.out_xlsx)
    out_md = Path(args.out_md)
    out_json = Path(args.out_json)

    only = (
        {int(x) for x in args.only.split(",") if x.strip()}
        if args.only.strip()
        else set(range(1, 99))
    )

    print(f"[1/4] master 로드: {master_path}")
    db = get_master_db(str(master_path))
    titles = db.article_titles()

    print(f"[2/4] HWP 파싱: {hwp_path}")
    text = parse_hwp(hwp_path)
    body = extract_body_region(text)
    articles_2026 = split_into_articles(body)
    print(f"      → {len(articles_2026)}개 조문 본문 추출 완료")

    model = get_llm_model()
    print(f"[3/4] LLM 추출 (model={model}, dry_run={args.dry_run})")
    records: list[dict] = []
    targets = sorted(n for n in only if 1 <= n <= 98)
    for idx, n in enumerate(targets, 1):
        title_2025 = titles.get(n) or db.title(n)
        body_2026 = articles_2026[n]
        g_old = db.law(n)
        i_old = db.penalty(n)

        cache_key = llm_cache.make_key(
            _SYSTEM_PROMPT,
            _build_user_prompt(n, title_2025, body_2026, g_old, i_old),
            _build_tool_schema(),
            model,
        )
        cached = llm_cache.get(cache_key) is not None
        try:
            payload = llm_extract(n, title_2025, body_2026, g_old, i_old, dry_run=args.dry_run)
        except Exception as e:
            print(f"  [{idx}/{len(targets)}] 제{n}조 — 실패: {type(e).__name__}: {e}")
            continue
        g_new = _serialize_law(payload.get("law") or [])
        i_new = _serialize_penalty(payload.get("penalty") or {})
        source_quote = payload.get("source_quote") or ""
        rec = {
            "no": n,
            "title": title_2025,
            "g_old": g_old,
            "g_new": g_new,
            "i_old": i_old,
            "i_new": i_new,
            "g_old_norm": _norm(g_old),
            "g_new_norm": _norm(g_new),
            "i_old_norm": _norm(i_old),
            "i_new_norm": _norm(i_new),
            "source_quote": source_quote,
            "raw": payload,
            "cache_hit": cached,
        }
        records.append(rec)
        flag = "cache" if cached else "live"
        print(
            f"  [{idx}/{len(targets)}] 제{n}조 ({title_2025}) [{flag}] "
            f"G {'변경' if rec['g_old_norm'] != rec['g_new_norm'] else '동일'} / "
            f"I {'변경' if rec['i_old_norm'] != rec['i_new_norm'] else '동일'}"
        )

    if not records:
        print("처리된 조문이 없음 — 종료")
        return 1

    reverted = revert_regressions(records)
    if reverted:
        print(
            f"[보정] LLM이 2025의 실질 벌칙 항목을 '직접 적용 벌칙 없음'으로 축소한 "
            f"{len(reverted)}건을 2025 값으로 복원: {reverted}"
        )

    print(f"[4/4] 산출물 저장")
    write_raw_json(out_json, records)
    print(f"      raw json → {out_json}")
    write_diff_md(out_md, records)
    print(f"      diff md  → {out_md}")
    if len(records) == 98:
        write_xlsx(master_path, out_xlsx, records, model_name=model)
        print(f"      xlsx     → {out_xlsx}")
    else:
        print(f"      xlsx     → 부분 실행이라 미생성 (전체 98개 처리 후 생성)")

    g_changed = sum(1 for r in records if r["g_old_norm"] != r["g_new_norm"])
    i_changed = sum(1 for r in records if r["i_old_norm"] != r["i_new_norm"])
    print(f"\n요약: G 변경 {g_changed}건 / I 변경 {i_changed}건 / 총 {len(records)}건 처리")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
